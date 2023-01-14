import pandas as pd

import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Side, Border

from openpyxl import styles, formatting


# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
from openpyxl.worksheet.worksheet import Worksheet
class Excell:
    col_names = ['Anno', 'Mese', 'Categoria', 'Voce', 'Euro']
    df_conti_camerino_modified = pd.read_excel('conti_camerino_da_importare.xlsx', names=col_names)

    # Creo la colonna ['Entrate_Uscite'] e scrivo in automatico i volori
    df_conti_camerino_modified['Entrate_Uscite'] = df_conti_camerino_modified['Categoria'].apply \
        (lambda x: 'Entrate' if x == 'Vendite varie' or
                                x == 'Salute' or
                                x == 'Curia' or
                                x == 'Collette-Chiesa' or
                                x == 'Congrua' or
                                x == 'Interessi' or
                                x == 'Messe celebrate' or
                                x == 'Offerte' or
                                x == 'Pensioni' or
                                x == 'Predicazione' or
                                x == 'Servizi religiosi' or
                                x == 'Stipendi' or
                                x == 'Sussidi' or
                                x == 'Rimbosi' or
                                x == 'Vendite varie' or
                                x == 'Eccedenza Cassa'
        else 'Uscite')

    # Assegno alle colonne l'ordine desiderato
    df_conti_camerino_modified = df_conti_camerino_modified[['Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro']]
    df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].astype("category")
    df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].cat.set_categories(
        ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno", "luglio", "agosto", "settembre", "ottobre",
         "novembre", "dicembre"])



    # sheets dei 12 mesi
    list_df_conti_camerino_mese_entrate = [ 'df_conti_camerino_mese_entrate_gennaio',
                                             'df_conti_camerino_mese_entrate_febbraio',
                                             'df_conti_camerino_mese_entrate_marzo',
                                             'df_conti_camerino_mese_entrate_aprile',
                                             'df_conti_camerino_mese_entrate_maggio',
                                             'df_conti_camerino_mese_entrate_giugno',
                                             'df_conti_camerino_mese_entrate_luglio',
                                             'df_conti_camerino_mese_entrate_agosto',
                                             'df_conti_camerino_mese_entrate_settembre',
                                             'df_conti_camerino_mese_entrate_ottobre',
                                             'df_conti_camerino_mese_entrate_novembre',
                                             'df_conti_camerino_mese_entrate_dicembre'
                                             ]

    list_df_conti_camerino_mese_uscite = [ 'df_conti_camerino_mese_uscite_gennaio',
                                            'df_conti_camerino_mese_uscite_febbraio',
                                            'df_conti_camerino_mese_uscite_marzo',
                                            'df_conti_camerino_mese_uscite_aprile',
                                            'df_conti_camerino_mese_uscite_maggio',
                                            'df_conti_camerino_mese_uscite_giugno',
                                            'df_conti_camerino_mese_uscite_luglio',
                                            'df_conti_camerino_mese_uscite_agosto',
                                            'df_conti_camerino_mese_uscite_settembre',
                                            'df_conti_camerino_mese_uscite_ottobre',
                                            'df_conti_camerino_mese_uscite_novembre',
                                            'df_conti_camerino_mese_uscite_dicembre'
                                            ]

    list_mese = ['gennaio',
                 'febbraio',
                 'marzo',
                 'aprile',
                 'maggio',
                 'giugno',
                 'luglio',
                 'agosto',
                 'settembre',
                 'ottobre',
                 'novembre',
                 'dicembre'
                 ]


    #imposto anno
    anno = 2015
    i = 0
    for x in range (12):
            list_df_conti_camerino_mese_entrate[i]= df_conti_camerino_modified.loc[
                (df_conti_camerino_modified['Anno'] == anno) &
                (df_conti_camerino_modified['Mese'] == list_mese[i]) &
                (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]
            #print(list_df_conti_camerino_pivot_entrate[i].head())


            list_df_conti_camerino_mese_uscite[i] = df_conti_camerino_modified.loc[
                (df_conti_camerino_modified['Anno'] == anno) &
                (df_conti_camerino_modified['Mese'] == list_mese[i]) &
                (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]
            #print(list_df_conti_camerino_mese_uscite[i].head())

            i += 1


    list_df_conti_camerino_mese_entrate = [  list_df_conti_camerino_mese_entrate[0], #df_conti_camerino_pivot_entrate_gennaio
                                                list_df_conti_camerino_mese_entrate[1],
                                                list_df_conti_camerino_mese_entrate[2],
                                                list_df_conti_camerino_mese_entrate[3],
                                                list_df_conti_camerino_mese_entrate[4],
                                                list_df_conti_camerino_mese_entrate[5],
                                                list_df_conti_camerino_mese_entrate[6],
                                                list_df_conti_camerino_mese_entrate[7],
                                                list_df_conti_camerino_mese_entrate[8],
                                                list_df_conti_camerino_mese_entrate[9],
                                                list_df_conti_camerino_mese_entrate[10],
                                                list_df_conti_camerino_mese_entrate[11]
                                            ]

    list_df_conti_camerino_mese_uscite = [   list_df_conti_camerino_mese_uscite[0], #df_conti_camerino_pivot_uscite_gennaio
                                                list_df_conti_camerino_mese_uscite[1],
                                                list_df_conti_camerino_mese_uscite[2],
                                                list_df_conti_camerino_mese_uscite[3],
                                                list_df_conti_camerino_mese_uscite[4],
                                                list_df_conti_camerino_mese_uscite[5],
                                                list_df_conti_camerino_mese_uscite[6],
                                                list_df_conti_camerino_mese_uscite[7],
                                                list_df_conti_camerino_mese_uscite[8],
                                                list_df_conti_camerino_mese_uscite[9],
                                                list_df_conti_camerino_mese_uscite[10],
                                                list_df_conti_camerino_mese_uscite[11]
                                            ]

    #print(list_df_conti_camerino_mese_entrate[0].columns)



    def pivot_table_w_subtotals(df, values, indices, columns, aggfunc, fill_value):
        listOfTable = []
        for indexNumber in range(len(indices)):
            n = indexNumber+1
            if n == 1:
                table = pd.pivot_table(df,values=values,index=indices[:n],columns=columns,aggfunc=aggfunc,fill_value=fill_value)

            else:
                table = pd.pivot_table(df,values=values,index=indices[:n],columns=columns,aggfunc=aggfunc,fill_value=fill_value)
            table = table.reset_index()


            for column in indices[n:]:
                table[column] = ''

            listOfTable.append(table)

        concatTable = pd.concat(listOfTable).sort_index()
        concatTable = concatTable.set_index(keys=indices)
        return concatTable.sort_index(axis=0,ascending=True)




    ''' 
    fin qui vengono creati 3 dataframes             print(table.sample())
    
    
    Index(['Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro'], dtype='object')
                          Euro
    Entrate_Uscite            
    Entrate         102266.604
    All             102266.604
                                  Euro
    Entrate_Uscite Categoria          
    Entrate        Predicazione  150.0
                                       Euro
    Entrate_Uscite Categoria Voce          
    Entrate        Offerte   Museo  1325.26
    
    '''





    list_pivot_mese_entrate = ['pivot_gennaio_entrate',
                                  'pivot_febbraio_entrate',
                                  'pivot_marzo_entrate',
                                  'pivot_aprile_entrate',
                                  'pivot_maggio_entrate',
                                  'pivot_giugno_entrate',
                                  'pivot_luglio_entrate',
                                  'pivot_agosto_entrate',
                                  'pivot_settembre_entrate',
                                  'pivot_ottobre_entrate',
                                  'pivot_novembre_entrate',
                                  'pivot_dicembre_entrate'
                                  ]

    list_pivot_mese_uscite = ['pivot_gennaio_uscite',
                                  'pivot_febbraio_uscite',
                                  'pivot_marzo_uscite',
                                  'pivot_aprile_uscite',
                                  'pivot_maggio_uscite',
                                  'pivot_giugno_uscite',
                                  'pivot_luglio_uscite',
                                  'pivot_agosto_uscite',
                                  'pivot_settembre_uscite',
                                  'pivot_ottobre_uscite',
                                  'pivot_novembre_uscite',
                                  'pivot_dicembre_uscite'
                                  ]

    # pivot_table_w_subtotals
    #       (df=list_df_conti_camerino_mese_entrate[0],values='Euro',indices=['Entrate_Uscite', 'Categoria', 'Voce'],columns=[],aggfunc='sum',fill_value='')


    i=0
    for x in range(12):
        list_pivot_mese_entrate[i] = np.round(pivot_table_w_subtotals
                                     (df=list_df_conti_camerino_mese_entrate[i],
                                      values='Euro',
                                      indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                      columns=[],
                                      aggfunc='sum',
                                      fill_value=''), 2)


        list_pivot_mese_uscite[i] = np.round(pivot_table_w_subtotals
                                     (list_df_conti_camerino_mese_uscite[i],
                                      values='Euro',
                                      indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                      columns=[],
                                      aggfunc='sum',
                                      fill_value=''), 2)
        i +=1

    list_pivot_mese_entrate=[list_pivot_mese_entrate[0], #pivot gennaio entrate
                        list_pivot_mese_entrate[1],
                        list_pivot_mese_entrate[2],
                        list_pivot_mese_entrate[3],
                        list_pivot_mese_entrate[4],
                        list_pivot_mese_entrate[5],
                        list_pivot_mese_entrate[6],
                        list_pivot_mese_entrate[7],
                        list_pivot_mese_entrate[8],
                        list_pivot_mese_entrate[9],
                        list_pivot_mese_entrate[10],
                        list_pivot_mese_entrate[11]
                        ]

    list_pivot_mese_uscite= [list_pivot_mese_uscite[0],
                        list_pivot_mese_uscite[1],
                        list_pivot_mese_uscite[2],
                        list_pivot_mese_uscite[3],
                        list_pivot_mese_uscite[4],
                        list_pivot_mese_uscite[5],
                        list_pivot_mese_uscite[6],
                        list_pivot_mese_uscite[7],
                        list_pivot_mese_uscite[8],
                        list_pivot_mese_uscite[9],
                        list_pivot_mese_uscite[10],
                        list_pivot_mese_uscite[11]
                        ]



    ##############PRIMA PAGINA
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment


    # Creo il file 'conti_camerino_styled.xlsx'
    wb = Workbook()
    # La prima pagina 'Sheet' la chiamo 'Copertina_fronte'
    wb['Sheet'].title = ('Copertina_fronte')

    wb['Copertina_fronte'].merge_cells('A4:I4')
    wb['Copertina_fronte']['A4'] = 'Resoconto Amministrativo'
    wb['Copertina_fronte']['A4'].font = Font(name='Calibri',
                                    size=35,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='204ac8') #blu royal
    wb['Copertina_fronte']['A4'].alignment = Alignment(horizontal='center')

    wb['Copertina_fronte']['A7'] = 'Fraternità di .....'
    wb['Copertina_fronte']['A7'].font = Font(name='Calibri',
                                    size=25,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='204ac8')

    wb['Copertina_fronte']['A10'] = 'Anno.....'
    wb['Copertina_fronte']['A10'].font = Font(name='Calibri',
                                    size=25,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='204ac8')



    # Inserisco immagine bilancia
    img = openpyxl.drawing.image.Image('bilancia.png')
    img.anchor = 'B13'
    wb['Copertina_fronte'].add_image(img)
    wb.save('conti_camerino_styled.xlsx')


    ##################################

    # Con ExcelWriter di pandas metto insieme il pivot delle entrate ed il pivot delle uscite
    list_mese = ['gennaio',
                 'febbraio',
                 'marzo',
                 'aprile',
                 'maggio',
                 'giugno',
                 'luglio',
                 'agosto',
                 'settembre',
                 'ottobre',
                 'novembre',
                 'dicembre'
                 ]


    # sheets dei 12 mesi
    list_ws_mese = ['ws_gennaio',
                    'ws_febbraio',
                    'ws_marzo',
                    'ws_aprile',
                    'ws_maggio',
                    'ws_giugno',
                    'ws_luglio',
                    'ws_agosto',
                    'ws_settembre',
                    'ws_ottobre',
                    'ws_novembre',
                    'ws_dicembre'
                    ]

    c = 0 # contatore

    for x in range (12):
        with pd.ExcelWriter('conti_camerino_styled.xlsx',
                            mode="a",
                            engine="openpyxl",
                            if_sheet_exists="overlay",
                            ) as writer:
            list_pivot_mese_entrate[c].to_excel(writer, sheet_name=list_mese[c], startrow=5)
            list_pivot_mese_uscite[c].to_excel(writer, sheet_name=list_mese[c], startrow=(len(list_pivot_mese_entrate[c] ) + 10))

        # leggo il file "conti_camerino_styled.xlsx"
        wb = load_workbook(filename="conti_camerino_styled.xlsx")
        #creo i 12 sheet
        list_ws_mese[c]= wb[list_mese[c]]

        c += 1



    #sheets dei 12 mesi
    list_ws_mese = [wb[list_mese[0]],  #ws_gennaio,
                    wb[list_mese[1]],  #ws_febbraio,
                    wb[list_mese[2]],  #ws_marzo,
                    wb[list_mese[3]],  #ws_aprile,
                    wb[list_mese[4]],  #ws_maggio,
                    wb[list_mese[5]],  #ws_giugno,
                    wb[list_mese[6]],  #ws_luglio,
                    wb[list_mese[7]],  #ws_agosto,
                    wb[list_mese[8]],  #ws_settembre,
                    wb[list_mese[9]],  #ws_ottobre,
                    wb[list_mese[10]],  #ws_novembre,
                    wb[list_mese[11]],  #ws_dicembre
                    ]


    ################# APPLICO STILE



    # Colonna D :Formattazione degli euro in valuta euro
    for sheet in list_ws_mese:
        for row in sheet[7:sheet.max_row]:  # skip the header
            #print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
            cell = row[3]  # il quarto valore della tuple
            # print (cell)# <Cell 'multiple'.D7>
            cell.number_format = '#,##0.00 €'
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)

    # Aggiungo la scritta 'Totale =' alla tabella pivot davanti ai subtotali
    for sheet in list_ws_mese:
        for row in sheet[7:sheet.max_row]:
            for cell in sheet['B']: #per ogni casella della colonna B
                if cell.value is not None:
                    # ossia se la casella nella colonna B non è vuota
                    # Assegnale uno stile
                    cell.font = Font(name='Calibri',
                                    size=15,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='a81a1a')
                    # Assegna uno stile anche alla cell accanto corrispondente
                    sheet.cell(row=cell.offset(row=0, column=0).row, column=3,
                                        value=f"Totale {cell.value} =")
                    sheet.cell(row=cell.offset(row=0, column=0).row, column=4).font\
                                        = Font(size=15, color='a81a1a', bold=True)
                    sheet.cell(row=cell.offset(row=0, column=0).row, column=4).number_format\
                                        = '#,##0.00 €'
                    sheet.cell(row=cell.offset(row=0, column=0).row, column=4).alignment\
                                        = Alignment(horizontal="left")



    # Larghezza fissa colonne
    i = 0  # contatore
    # set the height of the first row in each sheet
    for sheet in list_ws_mese:
        sheet.row_dimensions[1].height = 70

        # set the width of the column
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15

        # merge cells
        sheet.merge_cells('A1:D1')

        # scrivo nella cella 'A1'
        sheet['A1'].value = list_mese[i]
        i += 1

        # Formattazione cella
        sheet['A1'].font = Font(name='Calibri',
                                size=25,
                                bold=True,
                                italic=True,
                                vertAlign='none',
                                underline='single',
                                strike=False,
                                color='a81a1a')

        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Colonna D :Formattazione degli euro in valuta euro
    #for sheet in list_ws_mese:
        # for row in sheet[7:sheet.max_row]:  # skip the header
        #     # print(row) #(<Cell 'multiple'.A7>, <Cell 'multiple'.B7>, <Cell 'multiple'.C7>, <Cell 'multiple'.D7>)
        #     cell = row[3]  # il quarto valore della tuple
        #     # print (cell)# <Cell 'multiple'.D7>
        #     cell.number_format = '#,##0.00€'
        #     cell.alignment = Alignment(horizontal="right")
        #     cell.font = Font(bold=True)

        # Colonna C: Allineamento
        for row in sheet[7:sheet.max_row]:  # skip the header
            cell = row[2]  # il terzo valore della tuple
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Colonna D: Allineamento
        for row in sheet[7:sheet.max_row]:  # skip the header
            cell = row[1]  # il secondo valore della tuple
            cell.alignment = Alignment(horizontal="center", vertical="center")



        # Formattazione headers
        list = []

        for row in sheet.rows:
            for cell in row:
                if (cell.value == ("Categoria") or
                    cell.value == ("Entrate") or
                    cell.value == ("Euro") or
                    cell.value == ("Uscite") or
                    cell.value == ("Voce")):
                    list.append(cell)
        for cell in list:
            cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in sheet.rows:
            for cell in row:
                if (cell.value == ("Totale Categoria =")):

                        cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")


        # Formattazione ' Euro accanto a 'Totale categoria='
        from openpyxl.styles import PatternFill, GradientFill
        double = Side(border_style="double", color="4617F1")
        i = 0
        for row in sheet.rows:
            for cell in row:
                if (cell.value == ('Totale Categoria =')):
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4,
                                value=(list_df_conti_camerino_mese_entrate[i]['Euro']).sum(numeric_only=True))
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).font \
                                = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).number_format \
                                = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).alignment \
                                = Alignment(horizontal="left")
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).fill\
                            = PatternFill('solid', fgColor='d1d22e')
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).border\
                            = Border(bottom=double, top=double, left=double, right=double)

                        i += 1
        # Rendi 'invisibile il testo"Entrate_Uscite"
        list = []

        # Formattazione "Entrate_Uscite", piccolo per non essere visto

        for row in sheet.rows:
            for cell in row:
                if cell.value == ("Entrate_Uscite"):
                    list.append(cell)
        for cell in list:
            cell.font = Font(size=1)

        # for cell in list:
        #     cell.font = Font(size=1)
        #     print(cell)
        #     print(cell.coordinate)
        #     print(cell.row)
        #     print(cell.column)

        # Formattazione Euro somma totale
        # list = []
        #
        # for row in sheet.rows:
        #     for cell in row:
        #         if (cell.value == ("TOTALE_Entrate") or
        #             cell.value == ("TOTALE_Uscite")):
        #             list.append(cell)
        # for cell in list:
        #     sheet.cell(cell.row, column=4).font = Font(size=15, color='a81a1a', bold=True)

        # Creo tabella SAlDO  ws['A4'] = 4
    #            for sheet in sheetsLits:
        #         worksheet = workbook[sheet]
        #         FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)





        # sheets dei 12 mesi


    # list_df_conti_entrate = [list_df_conti_camerino_pivot_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
    #                           list_df_conti_camerino_pivot_entrate[1],
    #                           list_df_conti_camerino_pivot_entrate[2],
    #                           list_df_conti_camerino_pivot_entrate[3],
    #                           list_df_conti_camerino_pivot_entrate[4],
    #                           list_df_conti_camerino_pivot_entrate[5],
    #                           list_df_conti_camerino_pivot_entrate[6],
    #                           list_df_conti_camerino_pivot_entrate[7],
    #                           list_df_conti_camerino_pivot_entrate[8],
    #                           list_df_conti_camerino_pivot_entrate[9],
    #                           list_df_conti_camerino_pivot_entrate[10],
    #                           list_df_conti_camerino_pivot_entrate[11]
    #                           ]
    #
    # list_df_conti_uscite = [list_df_conti_camerino_pivot_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
    #                          list_df_conti_camerino_pivot_uscite[1],
    #                          list_df_conti_camerino_pivot_uscite[2],
    #                          list_df_conti_camerino_pivot_uscite[3],
    #                          list_df_conti_camerino_pivot_uscite[4],
    #                          list_df_conti_camerino_pivot_uscite[5],
    #                          list_df_conti_camerino_pivot_uscite[6],
    #                          list_df_conti_camerino_pivot_uscite[7],
    #                          list_df_conti_camerino_pivot_uscite[8],
    #                          list_df_conti_camerino_pivot_uscite[9],
    #                          list_df_conti_camerino_pivot_uscite[10],
    #                          list_df_conti_camerino_pivot_uscite[11]
    #                          ]



        # Riportare  entrate del mese in ogni foglio
    # i = 0  # contatore list_mese
    # for sheet in list_ws_mese:
    #     for row in sheet.rows:
    #         for cell in row:
    #             if (cell.value == ("TOTALE_Uscite")):
                    # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                    # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite


        # Riportare  entrate del mese in ogni foglio

        # for row in sheet.rows:
        #     for cell in row:
        #         if (cell.value == ("TOTALE_Uscite")):
                    # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                    # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite

        # for row in sheet.rows:
        #     for cell in row:
        #         if (cell.value == ("TOTALE_Uscite")):
                    # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                    # print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite


                    #i+= 1
    ##########

    ##################################################################################################################################
    #for sheet in list_ws_mese:
    #                last_cell_coordiate = 'C' + str(sheet.max_row)
    #                 print(last_cell_coordiate)
    #                 #attraverso le coordinate risalgo alla cella di excel
    #                 cell=sheet[last_cell_coordiate]
    #                 print(cell.value)
    #
    #                 sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font \
    #                       = Font(size=15, color='000000', bold=True)
    # saldo per tutti i mesi

    #imposto saldo iniziale
    saldo = 200_000
    #imposto un contatore
    e = 0
    i=0
    #memorizzo in una lista tutti i saldi dall'inizio alla fine dell'anno
    #le entrete
    #le uscite
    list_saldo_iniale_finale_anno=[saldo]
    list_entrate_mesi = []
    list_uscite_mesi = []
    #saldo iniziale
    for sheet in list_ws_mese:
                    #coordinate dell'ultima cella della colonna A di ogni foglio

                    last_cell_coordiate = 'C' + str(sheet.max_row)
                    print(last_cell_coordiate)
                    #attraverso le coordinate risalgo alla cella di excel
                    cell=sheet[last_cell_coordiate]
                    print(cell.value)

                    #cell = sheet.cell(row=1, column=1)
                    # last_cell = sheet[last_cell]
                    #
                    # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                    #
                    # #print(cell) # stampa ultima cella colonna A
                    sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font \
                          = Font(size=15, color='000000', bold=True)
                    sheet.cell(row=cell.offset(row=5, column=0).row, column=2).alignment = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=7, column=0).row, column=2, value='ENTRATE del mese').font \
                        = Font(size=15, color='000000', bold=True)
                    sheet.cell(row=cell.offset(row=7, column=0).row, column=2).alignment = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=2, value='USCITE del mese').font \
                        = Font(size=15, color='000000', bold=True)
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=2).alignment = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=2, value='DIS/AVANZO del mese').font \
                        = Font(size=15, color='000000', bold=True)
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=2).alignment = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=2, value='SALDO del mese corrente').font \
                        = Font(size=15, color='000000', bold=True)
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=2).alignment = Alignment(horizontal="left")

                    sheet.cell  (   row=cell.offset(row=5, column=0).row, column=4,
                                    value=  (
                                                    saldo
                                            )
                                )

                    # mi calcolo il saldo finale e la assegno alla variabile saldo
                    saldo = (saldo +
                             list_df_conti_camerino_mese_entrate[e]['Euro'].sum(numeric_only=True) -
                             list_df_conti_camerino_mese_uscite[e]['Euro'].sum(numeric_only=True)
                             )



                    #print(list_saldo_iniale_finale_anno)
                    sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(size=15, bold=True)
                    sheet.cell(row=cell.offset(row=5, column=0).row, column=4).number_format = '#,##0.00€'
                    sheet.cell(row=cell.offset(row=5, column=0).row, column=4).alignment = Alignment(horizontal="right")

                    if sheet.cell(row=cell.offset(row=5, column=0).row, column=4).value > 0:
                            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                            bold=True)
                    else:
                            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                            bold=True)

                    sheet.cell(row=cell.offset(row=7, column=0).row, column=4,
                               value=list_df_conti_camerino_mese_entrate[i]['Euro'].sum(numeric_only=True))


                    sheet.cell(row=cell.offset(row=7, column=0).row, column=4).font = Font(size=15, color='000000',
                                                                                           bold=True)
                    sheet.cell(row=cell.offset(row=7, column=0).row, column=4).number_format = '#,##0.00€'
                    sheet.cell(row=cell.offset(row=7, column=0).row, column=4).alignment = Alignment(horizontal="right")
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=4,
                               value=list_df_conti_camerino_mese_uscite[i]['Euro'].sum(numeric_only=True))
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=4).font = Font(size=15, color='a81a1a',
                                                                                           bold=True)
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=4).number_format = '-#,##0.00€'
                    sheet.cell(row=cell.offset(row=9, column=0).row, column=4).alignment = Alignment(horizontal="right")
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=4,
                               value=((list_df_conti_camerino_mese_entrate[i]['Euro']).sum(numeric_only=True) -
                                      (list_df_conti_camerino_mese_uscite[i]['Euro']).sum(numeric_only=True)
                                      )
                               )
                    # sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(size=15,
                    #                                                                       bold=True)
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=4).number_format = '#,##0.00€'
                    sheet.cell(row=cell.offset(row=11, column=0).row, column=4).alignment = Alignment(horizontal="right")

                    if (sheet.cell(row=cell.offset(row=11, column=0).row, column=4).value) > 0:
                        sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                                bold=True)
                    else:
                        sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='a81a1a', size=15,

                                                                                                bold=True)
                    #queste liste mi servono per il grafico
                    list_saldo_iniale_finale_anno.append(saldo)
                    list_entrate_mesi.append(list_df_conti_camerino_mese_entrate[i]['Euro'].sum(numeric_only=True))
                    list_uscite_mesi.append(list_df_conti_camerino_mese_uscite[i]['Euro'].sum(numeric_only=True))

                    # print(list_saldo_iniale_finale_anno)
                    # print(list_entrate_mesi)
                    # print(list_uscite_mesi)

                    i += 1
    # Saldo finale
    # for row in sheet:
    #         for cell in row:
    #             if (cell.value == ("TOTALE_Uscite")):

                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4,
                                value=  (
                                            saldo
                                        )
                                )

                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(size=15, bold=True)
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4).number_format = '#,## 0.00€'
                    sheet.cell(row=cell.offset(row=13, column=0).row, column=4).alignment = Alignment(horizontal="right")

                    if sheet.cell(row=cell.offset(row=13, column=0).row, column=4).value > 0:
                        sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='000000', size=15, bold=True)
                    else:
                        sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='a81a1a', size=15, bold=True)

    ###################### tabellone entrate
    import openpyxl
    from openpyxl.worksheet import page
    from openpyxl.utils.dataframe import dataframe_to_rows

    #Creo un nuovo foglio
    ws_tab_entrate = wb.create_sheet('Tab_Entrate')
    ws_tab_entrate.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)
    #ws_tab_entrate.print_area = 'A1:Z1'

    ws_tab_entrate['A1']='Tabellone Entrate'
    ws_tab_entrate.row_dimensions[1].height = 70
    ws_tab_entrate['A1'].font = Font(name='Calibri', size=80, color='a81a1a', bold=True)
    ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_tab_entrate.merge_cells('A1:N1')





     # set the width of the column
    ws_tab_entrate.column_dimensions['A'].width = 14
    ws_tab_entrate.column_dimensions['B'].width = 8
    ws_tab_entrate.column_dimensions['C'].width = 8
    ws_tab_entrate.column_dimensions['D'].width = 8
    ws_tab_entrate.column_dimensions['E'].width = 8
    ws_tab_entrate.column_dimensions['F'].width = 8
    ws_tab_entrate.column_dimensions['G'].width = 8
    ws_tab_entrate.column_dimensions['H'].width = 8
    ws_tab_entrate.column_dimensions['I'].width = 8
    ws_tab_entrate.column_dimensions['L'].width = 9
    ws_tab_entrate.column_dimensions['M'].width = 8
    ws_tab_entrate.column_dimensions['N'].width = 14

    i=2
    for x in range(2, 20):
        ws_tab_entrate.row_dimensions[i].height = 20
        i +=1



    anno = 2015
    df_conti_camerino_TOT_entrate= df_conti_camerino_modified.loc[
                (df_conti_camerino_modified['Anno'] == anno) &
                (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]
    #print(df_conti_camerino_TOT_entrate.head(40))



    pivot_conti_camerino_TOT_entrate = np.round(pd.pivot_table
                                     (df_conti_camerino_TOT_entrate,
                                      values='Euro',
                                      index=['Categoria'],
                                      columns='Mese',
                                      aggfunc='sum',
                                      margins=True,
                                      margins_name='TOTALE_Entrate',
                                      fill_value=0), 2)

    #print(pivot_conti_camerino_TOT_entrate.head())

    for r in dataframe_to_rows(pivot_conti_camerino_TOT_entrate, index=True, header=True):
        ws_tab_entrate.append(r)

    for cell in ws_tab_entrate['A'] + ws_tab_entrate[2]:
        cell.style = 'Pandas'

    ws_tab_entrate['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
    ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")

    TOTALE_ENTRATE = round(df_conti_camerino_TOT_entrate['Euro'].sum(),2)



    for row in ws_tab_entrate.rows:
            for cell in row:
                if  cell.value == TOTALE_ENTRATE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,## 0.00€'





            # list_df_conti_camerino_mese_uscite[i] = df_conti_camerino_modified.loc[
            #     (df_conti_camerino_modified['Anno'] == anno) &
            #     (df_conti_camerino_modified['Mese'] == list_mese[i]) &
            #     (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]
            # #print(list_df_conti_camerino_mese_uscite[i].head())
            #
            # i += 1


    ###################### tabellone uscite
    #Creo un nuovo foglio
    ws_tab_uscite = wb.create_sheet('Tab_Uscite')

    ws_tab_uscite.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)

    ws_tab_uscite['A1']='Tabellone Uscite'
    ws_tab_uscite.merge_cells('A1:N1')
    ws_tab_uscite.row_dimensions[1].height = 45


    df_conti_camerino_TOT_uscite= df_conti_camerino_modified.loc[
                (df_conti_camerino_modified['Anno'] == anno) &
                (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

    #print(df_conti_camerino_TOT_uscite.head(40))
    pivot_conti_camerino_TOT_uscite= np.round(pd.pivot_table
                                     (df_conti_camerino_TOT_uscite,
                                      values='Euro',
                                      #index=['Entrate_Uscite', 'Categoria', 'Voce'],
                                      index=['Categoria'],
                                      columns='Mese',
                                      aggfunc='sum',
                                      margins=True,
                                      margins_name='TOTALE_Uscite',
                                      fill_value=0), 2)

    print(pivot_conti_camerino_TOT_uscite.head())
    for r in dataframe_to_rows(pivot_conti_camerino_TOT_uscite, index=True, header=True):
        ws_tab_uscite.append(r)

    for cell in ws_tab_uscite['A'] + ws_tab_uscite[2]:
        cell.style = 'Pandas'

    # set the width of the column
    ws_tab_uscite.column_dimensions['A'].width = 16
    ws_tab_uscite.column_dimensions['B'].width = 7
    ws_tab_uscite.column_dimensions['C'].width = 7
    ws_tab_uscite.column_dimensions['D'].width = 7
    ws_tab_uscite.column_dimensions['E'].width = 7
    ws_tab_uscite.column_dimensions['F'].width = 7
    ws_tab_uscite.column_dimensions['G'].width = 7
    ws_tab_uscite.column_dimensions['H'].width = 7
    ws_tab_uscite.column_dimensions['I'].width = 7
    ws_tab_uscite.column_dimensions['L'].width = 9
    ws_tab_uscite.column_dimensions['M'].width = 7
    ws_tab_uscite.column_dimensions['N'].width = 16

    i=2
    for x in range(2, 20):
        ws_tab_uscite.row_dimensions[i].height = 17
        i +=1


    ws_tab_uscite['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
    ws_tab_uscite['A1'].alignment = Alignment(horizontal="center", vertical="center")

    TOTALE_USCITE = round(df_conti_camerino_TOT_uscite['Euro'].sum(),2)




    for row in ws_tab_uscite.rows:
            for cell in row:
                if  cell.value == TOTALE_USCITE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,## 0.00€'

    ######################  grafico

    from openpyxl.chart import Reference, LineChart

    #Creo un nuovo foglio
    ws_saldo_riepilogo = wb.create_sheet('Saldo_riepilogo')
    ws_saldo_riepilogo['A1']='Tabella Entrate - Uscite - Saldo di ogni mese '

    #Formattazione
    ws_saldo_riepilogo.row_dimensions[1].height = 100
    ws_saldo_riepilogo.merge_cells('A1:E1')





     # set the width of the column
    ws_saldo_riepilogo.column_dimensions['A'].width = 15
    ws_saldo_riepilogo.column_dimensions['B'].width = 17
    ws_saldo_riepilogo.column_dimensions['C'].width = 17
    ws_saldo_riepilogo.column_dimensions['D'].width = 17
    ws_saldo_riepilogo.column_dimensions['E'].width = 17


    #queste liste mi servono per il grafico

    list_headers= ['mese', 'saldo_iniziale','entrate_mese', 'uscite_mese', 'saldo_finale']
    list_saldo_iniziale_anno = list_saldo_iniale_finale_anno[:-1]
    # list_entrate_mesi
    # list_uscite_mesi
    list_saldo_finale_anno = list_saldo_iniale_finale_anno[1:]


    i = 0
    ws_saldo_riepilogo.append(list_headers)
    for mese in range(1,13):
        mese_saldo_grafico = [list_mese[i], list_saldo_iniziale_anno[i], list_entrate_mesi[i], list_uscite_mesi[i], list_saldo_finale_anno[i]]

        ws_saldo_riepilogo.append(mese_saldo_grafico)
        i +=1


    list = []
    for row in ws_saldo_riepilogo.rows:
            for cell in row:


                if (cell.value == ("saldo_iniziale") or
                    cell.value == ("entrate_mese") or
                    cell.value == ("uscite_mese") or
                    cell.value == ("saldo_finale") or
                    cell.value == ("gennaio") or
                    cell.value == ("febbraio") or
                    cell.value == ("marzo") or
                    cell.value == ("aprile") or
                    cell.value == ("maggio") or
                    cell.value == ("giugno") or
                    cell.value == ("luglio") or
                    cell.value == ("agosto") or
                    cell.value == ("settembre") or
                    cell.value == ("ottobre") or
                    cell.value == ("novembre") or
                    cell.value == ("dicembre")
                    ):
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for cell in row:
                if  cell.value == ("mese"):
                    cell.font = Font(size=1)

    for row in ws_saldo_riepilogo.iter_rows(min_row=3, min_col=2, max_row=14, max_col=5):
            for cell in row:
                cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,## 0.00 €'
                if (int(cell.value) > 0):
                    cell.font = Font(color='000000')
                else:
                    cell.font = Font(color='a81a1a')





    ws_saldo_riepilogo['A1'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
    ws_saldo_riepilogo['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_saldo_riepilogo['B3'].border\
                            = Border(bottom=double, top=double, left=double, right=double)
    ws_saldo_riepilogo['B3'].fill\
                            = PatternFill('solid', fgColor='d1d22e')
    ws_saldo_riepilogo['E14'].border\
                            = Border(bottom=double, top=double, left=double, right=double)
    ws_saldo_riepilogo['E14'].fill\
                            = PatternFill('solid', fgColor='d1d22e')

    data = Reference(ws_saldo_riepilogo, min_col=3, min_row=2, max_col=5, max_row=14)
    titles = Reference(ws_saldo_riepilogo, min_row=3, max_row=14, min_col=1)

    chart = LineChart()
    chart.title = "Bilancio"
    chart.style = 12

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(titles)
    chart.x_axis.title = 'Mesi'
    chart.y_axis.title = 'Euro'


    ws_saldo_riepilogo.add_chart(chart, "A21")
    #######################PAGINA CONCLUSIVA
    #format
    from openpyxl.styles import NamedStyle, Font, Border, Side
    #HIGHLIGHT
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name='Calibri', size=15, color='000000', bold=True)
    double = Side(border_style="double", color="4617F1")
    #highlight.border = Border(bottom=double, top=double, left=double, right=double)
    highlight.fill = PatternFill('solid', fgColor='d1d22e')
    highlight.alignment = Alignment(horizontal="right", vertical="center")
    highlight.number_format = '#,## 0.00 €'
    wb.add_named_style(highlight)

    ############BLACK
    black = NamedStyle(name="black")
    black.font = Font(name='Calibri', size=15, color='000000', bold=True)
    black.alignment = Alignment(horizontal="right", vertical="center")
    wb.add_named_style(black)



    import openpyxl
    from openpyxl import load_workbook
    ws_fine = wb.create_sheet('Fine')
    ws_fine['A1']='Bilancio Anno ...'
    ws_fine.merge_cells('A1:I1')
    ws_fine.row_dimensions[1].height = 45
    ws_fine['A1'].font = Font(name='Calibri', size=35, color='a81a1a', bold=True)
    ws_fine['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws_fine['A7']='RIEPILOGO'
    ws_fine.merge_cells('A7:C7')

    ws_fine['A10']='SALDO iniziale'
    ws_fine.merge_cells('A10:C10')
    ws_fine['E10']=int(200_000)
    ws_fine.merge_cells('E10:G10')
    ws_fine['E10'].style = 'highlight'

    ws_fine['A13']='TOTALE Entrate'
    ws_fine.merge_cells('A13:C13')
    ws_fine['E13']=TOTALE_ENTRATE
    ws_fine.merge_cells('E13:G13')
    ws_fine['E13'].style = 'highlight'

    ws_fine['A16']='TOTALE Uscite'
    ws_fine.merge_cells('A16:C16')
    ws_fine['E16']=TOTALE_USCITE
    ws_fine.merge_cells('E16:G16')
    ws_fine['E16'].style = 'highlight'

    ws_fine['A19']='DIS/AVANZO'
    ws_fine.merge_cells('A19:C19')
    ws_fine['E19']=(TOTALE_ENTRATE - TOTALE_USCITE)
    ws_fine.merge_cells('E19:G19')
    ws_fine['E19'].style = 'highlight'

    ws_fine['A22']='SALDO Finale'
    ws_fine.merge_cells('A22:C22')
    ws_fine['E22']=saldo
    ws_fine.merge_cells('E22:G22')
    ws_fine['E22'].style = 'highlight'

    ws_fine['A7'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
    ws_fine['A7'].alignment = Alignment(horizontal="center", vertical="center")
    ws_fine['A10'].font = Font(name='Calibri', size=15, bold=True)
    ws_fine['A10'].alignment = Alignment(horizontal="right", vertical="center")
    ws_fine['A13'].font = Font(name='Calibri', size=15, bold=True)
    ws_fine['A13'].alignment = Alignment(horizontal="right", vertical="center")
    ws_fine['A16'].font = Font(name='Calibri', size=15, bold=True)
    ws_fine['A16'].alignment = Alignment(horizontal="right", vertical="center")
    ws_fine['A19'].font = Font(name='Calibri', size=15, bold=True)
    ws_fine['A19'].alignment = Alignment(horizontal="right", vertical="center")
    ws_fine['A22'].font = Font(name='Calibri', size=15, bold=True)
    ws_fine['A22'].alignment = Alignment(horizontal="right", vertical="center")


    ws_fine['A32']='Data'
    ws_fine['A32'].style = 'black'
    ws_fine.merge_cells('A32:B32')
    ws_fine['D29']='Guardiano'
    ws_fine['D29'].style = 'black'
    ws_fine.merge_cells('D29:I29')
    ws_fine['d35']='Vicario'
    ws_fine['D35'].style = 'black'
    ws_fine.merge_cells('D35:I35')
    ws_fine['D41']='Economo'
    ws_fine['D41'].style = 'black'
    ws_fine.merge_cells('D41:I41')
    ws_fine['A38']='Timbro'
    ws_fine['A38'].style = 'black'
    ws_fine.merge_cells('A38:B38')




    # Salva
    wb.save("conti_camerino_styled.xlsx")
    print('ESEGUITO')

Excell()