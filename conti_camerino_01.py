
import pandas as pd

import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
col_names = ['Anno', 'Mese', 'Categoria', 'Voce', 'Euro']
df_conti_camerino_modified = pd.read_excel('conti_camerino_da_importare.xlsx', names=col_names)


# Creo la colonna ['Entrate_Uscite'] e scrivo in automatico i volori
df_conti_camerino_modified['Entrate_Uscite']= df_conti_camerino_modified['Categoria'].apply\
    (lambda x: 'Entrate'    if  x=='Vendite varie' or
                                x=='Salute' or
                                x=='Curia' or
                                x=='Collette-Chiesa' or
                                x=='Congrua' or
                                x=='Interessi' or
                                x=='Messe celebrate' or
                                x=='Offerte' or
                                x=='Pensioni' or
                                x=='Predicazione' or
                                x=='Servizi religiosi' or
                                x=='Stipendi' or
                                x=='Sussidi' or
                                x=='Rimbosi' or
                                x=='Vendite varie' or
                                x=='Eccedenza Cassa'
                            else 'Uscite')

# Assegno alle colonne l'ordine desiderato
df_conti_camerino_modified = df_conti_camerino_modified [['Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce','Euro']]
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].astype("category")
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].cat.set_categories(["gennaio","febbraio","marzo","aprile","maggio","giugno","luglio","agosto","settembre","ottobre","novembre","dicembre"])


#creo dataframe per le entrate di ogni mese di uno specifico anno
df_conti_camerino_pivot_entrate_gennaio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'gennaio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]


df_conti_camerino_pivot_entrate_febbraio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'febbraio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_marzo = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'marzo') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_aprile = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'aprile') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_maggio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'maggio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_giugno = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'giugno') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_luglio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'luglio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_agosto = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'agosto') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_settembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'settembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_ottobre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'ottobre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_novembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'novembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_entrate_dicembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'dicembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]


#creo dataframe per le uscite di ogni mese di uno specifico anno


df_conti_camerino_pivot_uscite_gennaio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'gennaio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_febbraio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'febbraio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_marzo = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'marzo') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_aprile = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'aprile') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_maggio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'maggio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_giugno = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'giugno') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_luglio = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'luglio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_agosto = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'agosto') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_settembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'settembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_ottobre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'ottobre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_novembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'novembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]

df_conti_camerino_pivot_uscite_dicembre = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'dicembre') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]


# Creo pivot per le entrate di ogni mese di uno specifico anno
pivot_gennaio_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_gennaio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_febbraio_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_febbraio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_marzo_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_marzo,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_aprile_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_aprile,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_maggio_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_maggio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_giugno_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_giugno,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_luglio_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_luglio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_agosto_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_agosto,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_settembre_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_settembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_ottobre_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_ottobre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_novembre_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_novembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

pivot_dicembre_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_entrate_dicembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Entrate',
                               fill_value=0),2)

# Creo pivot per le uscite di ogni mese di uno specifico anno
# round arrotonda alla seconda cifra decimale ',2)'


pivot_gennaio_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_gennaio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_febbraio_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_febbraio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_marzo_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_marzo,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_aprile_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_aprile,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_maggio_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_maggio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_giugno_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_giugno,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_luglio_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_luglio,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_agosto_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_agosto,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_settembre_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_settembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_ottobre_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_ottobre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_novembre_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_novembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)

pivot_dicembre_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_uscite_dicembre,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE_Uscite',
                               fill_value=0),2)


#Creo il file 'conti_camerino_styled.xlsx'
wb = Workbook()
wb['Sheet'].title = 'Copertina_fronte'
wb.save('conti_camerino_styled.xlsx')


# Con ExcelWriter di pandas metto insieme il pivot delle entrate ed il pivot delle uscite

with pd.ExcelWriter('conti_camerino_styled.xlsx',
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_gennaio_entrate.to_excel(writer, sheet_name="Gennaio", startrow=5)
                    pivot_gennaio_uscite.to_excel(writer, sheet_name="Gennaio", startrow=(len(pivot_gennaio_entrate)+10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_febbraio_entrate.to_excel(writer, sheet_name="Febbraio", startrow=5)
                    pivot_febbraio_uscite.to_excel(writer, sheet_name="Febbraio", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_marzo_entrate.to_excel(writer, sheet_name="Marzo", startrow=5)
                    pivot_marzo_uscite.to_excel(writer, sheet_name="Marzo", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_aprile_entrate.to_excel(writer, sheet_name="Aprile", startrow=5)
                    pivot_aprile_uscite.to_excel(writer, sheet_name="Aprile", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_maggio_entrate.to_excel(writer, sheet_name="Maggio", startrow=5)
                    pivot_maggio_uscite.to_excel(writer, sheet_name="Maggio", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_giugno_entrate.to_excel(writer, sheet_name="Giugno", startrow=5)
                    pivot_giugno_uscite.to_excel(writer, sheet_name="Giugno", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_luglio_entrate.to_excel(writer, sheet_name="Luglio", startrow=5)
                    pivot_luglio_uscite.to_excel(writer, sheet_name="Luglio", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_agosto_entrate.to_excel(writer, sheet_name="Agosto", startrow=5)
                    pivot_agosto_uscite.to_excel(writer, sheet_name="Agosto", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_settembre_entrate.to_excel(writer, sheet_name="Settembre", startrow=5)
                    pivot_settembre_uscite.to_excel(writer, sheet_name="Settembre", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_ottobre_entrate.to_excel(writer, sheet_name="Ottobre", startrow=5)
                    pivot_ottobre_uscite.to_excel(writer, sheet_name="Ottobre", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_novembre_entrate.to_excel(writer, sheet_name="Novembre", startrow=5)
                    pivot_novembre_uscite.to_excel(writer, sheet_name="Novembre", startrow=(len(pivot_gennaio_entrate) + 10))


with pd.ExcelWriter("conti_camerino_styled.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_dicembre_entrate.to_excel(writer, sheet_name="Dicembre", startrow=5)
                    pivot_dicembre_uscite.to_excel(writer, sheet_name="Dicembre", startrow=(len(pivot_gennaio_entrate) + 10))


# leggo il file "conti_camerino_styled.xlsx"
wb = load_workbook(filename = "conti_camerino_styled.xlsx")

# creo 12 sheet per i 12 mesi
ws_gennaio  = wb['Gennaio']
ws_febbraio = wb['Febbraio']
ws_marzo    = wb['Marzo']
ws_aprile   = wb['Aprile']
ws_maggio   = wb['Maggio']
ws_giugno   = wb['Giugno']
ws_luglio   = wb['Luglio']
ws_agosto   = wb['Agosto']
ws_settembre = wb['Settembre']
ws_ottobre  = wb['Ottobre']
ws_novembre = wb['Novembre']
ws_dicembre = wb['Dicembre']

# creo 2 liste: i fogli di excel ed i mesi dell'anno

#sheets dei 12 mesi
list_ws_mese = [ws_gennaio,
                ws_febbraio,
                ws_marzo,
                ws_aprile,
                ws_maggio,
                ws_giugno,
                ws_luglio,
                ws_agosto,
                ws_settembre,
                ws_ottobre,
                ws_novembre,
                ws_dicembre]

i = 0 # contatore list_mese
list_mese = ['Conto del mese di Gennaio',
             'Conto del mese di Febbraio',
             'Conto del mese di Marzo',
             'Conto del mese di Aprile',
             'Conto del mese di Maggio',
             'Conto del mese di Giugno',
             'Conto del mese di Luglio',
             'Conto del mese di Agosto',
             'Conto del mese di Settembre',
             'Conto del mese di Ottobre',
             'Conto del mese di Novembre',
             'Conto del mese di Dicembre',
             ]

# set the height of the first row in each sheet
for sheet in list_ws_mese:
    sheet.row_dimensions[1].height = 70


# set the width of the column
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15

#merge cells
    sheet.merge_cells('A1:D1')

    #scrivo nella cella 'A1'
    sheet['A1'].value = list_mese[i]
    i += 1

    # Formattazione cella
    sheet['A1'].font=Font(  name='Calibri',
                            size=25,
                            bold=True,
                            italic=True,
                            vertAlign='none',
                            underline='single',
                            strike=False,
                            color='a81a1a')

    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Colonna D :Formattazione degli euro in valuta euro
for sheet in list_ws_mese:
    for row in sheet[7:sheet.max_row]:  # skip the header
        #print(row) #(<Cell 'multiple'.A7>, <Cell 'multiple'.B7>, <Cell 'multiple'.C7>, <Cell 'multiple'.D7>)
        cell = row[3] #il quarto valore della tuple
        #print (cell)# <Cell 'multiple'.D7>
        cell.number_format = '#,##0.00€'
        cell.alignment = Alignment(horizontal="right")
        cell.font=Font(bold=True)

# Colonna C: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[2]  #il terzo valore della tuple
        cell.alignment = Alignment(horizontal="right")

# Colonna D: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[1]  #il secondo valore della tuple
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Formattazione headers
    list=[]

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("Categoria") or
                cell.value == ("Entrate") or
                cell.value == ("Euro") or
                cell.value == ("Uscite") or
                cell.value == ("Voce")):
                list.append(cell)
    for cell in list:
        cell.font = Font(size=15, color='a81a1a', bold=True)

# Formattazione 'TOTALE_Entrate' e 'TOTALE_Uscite'
    list = []

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ('TOTALE_Entrate') or
                cell.value == ('TOTALE_Uscite')):
                list.append(cell)
        for cell in list:
            cell.font = Font(size=12, color='a81a1a', bold=True)



# Rendi 'invisibile il testo"Entrate_Uscite"
    list=[]

    for row in sheet.rows:
        for cell in row:
            if cell.value == ("Entrate_Uscite"):
                list.append(cell)
    for cell in list:
        cell.font = Font(size=1)

    # for cell in list:
    #     cell.font = Font(size=1)
    #     print(cell)
    #     print(cell.coordinate)
    #     print(cell.row)
    #     print(cell.column)

# Formattazione Euro somma totale
    list=[]

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Entrate") or
                cell.value == ("TOTALE_Uscite")):
                list.append(cell)
    for cell in list:
        sheet.cell(cell.row, column=4).font = Font(size=15, color='a81a1a', bold=True)


#Creo tabella SAlDO
# text


    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("TOTALE_Uscite")):
                #print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
                #print(sheet.cell(row=cell.row, column=cell.column).value) # stampa 'TOTALE_Uscite
                sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font\
                    = Font(size=15, color='a81a1a', bold=True)
                sheet.cell(row=cell.offset(row=7, column=0).row, column=2, value='ENTRATE del mese').font\
                    = Font(size=15, color='a81a1a', bold=True)
                sheet.cell(row=cell.offset(row=9, column=0).row, column=2, value='USCITE del mese').font\
                    = Font(size=15, color='a81a1a', bold=True)
                sheet.cell(row=cell.offset(row=11, column=0).row, column=2, value='DIS/AVANZO del mese').font\
                    = Font(size=15, color='a81a1a', bold=True)
                sheet.cell(row=cell.offset(row=13, column=0).row, column=2, value='SALDO del mese corrente').font\
                    = Font(size=15, color='a81a1a', bold=True)


# from openpyxl import load_workbook
#
# wb = load_workbook("copies.xlsx")
# ws1 = wb.worksheets[0]
# ws2 = wb.worksheets[1]
#
# # Get the row
# for rows in ws1.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
#     # Get the cell
#     for cell in rows:
#         # offset the values for cells on ws2 with cell offset
#         ws2.cell(row=cell.offset(row=5, column=0).row, column=1, value=cell.value)
# wb.save("copies_copied.xlsx")


# Salva
wb.save("conti_camerino_styled.xlsx")

