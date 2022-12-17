from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import sqlite3
import pandas as pd

# Create a Workbook object
#wb = Workbook()


#wb = load_workbook('database_conti.xlsx')

# Create a database or connect to one that exists
conn = sqlite3.connect('database_conti')

# Create a cursor instance
c = conn.cursor()

query = "SELECT * FROM TABLE_Conti"  # query to collect recors
df = pd.read_sql(query, conn, index_col='ID')  # create dataframe
df.to_excel('database_conti.xlsx')  # create excel file

wb = load_workbook('database_conti.xlsx')

#Create a worksheet
ws = wb.active

#print a cell
print(ws['G6'].value)
print(f'{ws["B6"].value} mese di {ws["C6"].value}')

#print una intera colonna
colonna_E = ws['E']
for cell in colonna_E:
    print(cell.value)

# print una intera riga
print('######################')
riga_2 = ws['1']
for cell in riga_2:
    print(cell.value)



# Commit changes
conn.commit()

# Close our connection
conn.close()
