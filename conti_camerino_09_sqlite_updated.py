import pandas as pd
import numpy as np
import sqlite3

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Side, Border
from openpyxl.styles import PatternFill





#CONNESSIONE A SQLITE3
conn = sqlite3.connect('database_conti')
cur = conn.cursor()

#MI ASSIOCURO DI ESSERE CONNESSO
print(conn)
print('Sei connesso al database_conti')

#CREO DATAFRAME df_database_conti
df_database_conti = pd.read_sql("select * from TABLE_Conti", conn)
print(df_database_conti.info(verbose=True))


#COMMIT e CLOSE
conn.commit()
conn.close()


# LIST TITOLI DATAFRAMES 12 MESI
list_df_conti_mese_entrate =   ['df_database_conti_entrate_gennaio',
                                         'df_database_conti_entrate_febbraio',
                                         'df_database_conti_entrate_marzo',
                                         'df_database_conti_aprile',
                                         'df_database_conti_maggio',
                                         'df_database_conti_giugno',
                                         'df_database_conti_luglio',
                                         'df_database_conti_agosto',
                                         'df_database_conti_settembre',
                                         'df_database_conti_ottobre',
                                         'df_database_conti_novembre',
                                         'df_database_conti_dicembre'
                                         ]

list_df_conti_mese_uscite = ['df_database_conti_uscite_gennaio',
                                      'df_database_conti_uscite_febbraio',
                                      'df_database_conti_uscite_marzo',
                                      'df_database_conti_uscite_aprile',
                                      'df_database_conti_uscite_maggio',
                                      'df_database_conti_uscite_giugno',
                                      'df_database_conti_uscite_luglio',
                                      'df_database_conti_uscite_agosto',
                                      'df_database_conti_uscite_settembre',
                                      'df_database_conti_uscite_ottobre',
                                      'df_database_conti_uscite_novembre',
                                      'df_database_conti_uscite_dicembre'
                                      ]

list_mese = ['gennaio',
             'febbraio',
             'marzo',
             'aprile',
             'maggio',
             'giugno',
             'luglio',
             'agosto',
             'settembre',
             'ottobre',
             'novembre',
             'dicembre'
             ]

#imposto anno
anno = 2012
i = 0
for x in range (12):
        list_df_conti_mese_entrate[i]= df_database_conti.loc[
            (df_database_conti['Anno'] == anno) &
            (df_database_conti['Mese'] == list_mese[i]) &
            (df_database_conti['Entrate_Uscite'] == 'Entrate')]
        #print(list_df_conti_camerino_pivot_entrate[i].head())


        list_df_conti_mese_uscite[i] = df_database_conti.loc[
            (df_database_conti['Anno'] == anno) &
            (df_database_conti['Mese'] == list_mese[i]) &
            (df_database_conti['Entrate_Uscite'] == 'Uscite')]
        #print(list_df_conti_camerino_mese_uscite[i].head())

        i += 1


list_df_conti_mese_entrate = [  list_df_conti_mese_entrate[0], #df_conti_camerino_pivot_entrate_gennaio
                                            list_df_conti_mese_entrate[1],
                                            list_df_conti_mese_entrate[2],
                                            list_df_conti_mese_entrate[3],
                                            list_df_conti_mese_entrate[4],
                                            list_df_conti_mese_entrate[5],
                                            list_df_conti_mese_entrate[6],
                                            list_df_conti_mese_entrate[7],
                                            list_df_conti_mese_entrate[8],
                                            list_df_conti_mese_entrate[9],
                                            list_df_conti_mese_entrate[10],
                                            list_df_conti_mese_entrate[11]
                                        ]

list_df_conti_mese_uscite = [   list_df_conti_mese_uscite[0], #df_conti_camerino_pivot_uscite_gennaio
                                            list_df_conti_mese_uscite[1],
                                            list_df_conti_mese_uscite[2],
                                            list_df_conti_mese_uscite[3],
                                            list_df_conti_mese_uscite[4],
                                            list_df_conti_mese_uscite[5],
                                            list_df_conti_mese_uscite[6],
                                            list_df_conti_mese_uscite[7],
                                            list_df_conti_mese_uscite[8],
                                            list_df_conti_mese_uscite[9],
                                            list_df_conti_mese_uscite[10],
                                            list_df_conti_mese_uscite[11]
                                        ]

#print(list_df_conti_mese_entrate[0].columns)


def pivot_table_w_subtotals(df, values, indices, columns, aggfunc, fill_value):
    listOfTable = []
    for indexNumber in range(len(indices)):
        n = indexNumber+1
        if n == 1:
            table = pd.pivot_table(df,values=values,index=indices[:n],columns=columns,aggfunc=aggfunc,fill_value=fill_value)

        else:
            table = pd.pivot_table(df,values=values,index=indices[:n],columns=columns,aggfunc=aggfunc,fill_value=fill_value)
        table = table.reset_index()


        for column in indices[n:]:
            table[column] = ''

        listOfTable.append(table)

    concatTable = pd.concat(listOfTable).sort_index()
    concatTable = concatTable.set_index(keys=indices)
    return concatTable.sort_index(axis=0,ascending=True)


list_pivot_mese_entrate = ['pivot_gennaio_entrate',
                              'pivot_febbraio_entrate',
                              'pivot_marzo_entrate',
                              'pivot_aprile_entrate',
                              'pivot_maggio_entrate',
                              'pivot_giugno_entrate',
                              'pivot_luglio_entrate',
                              'pivot_agosto_entrate',
                              'pivot_settembre_entrate',
                              'pivot_ottobre_entrate',
                              'pivot_novembre_entrate',
                              'pivot_dicembre_entrate'
                              ]

list_pivot_mese_uscite = ['pivot_gennaio_uscite',
                              'pivot_febbraio_uscite',
                              'pivot_marzo_uscite',
                              'pivot_aprile_uscite',
                              'pivot_maggio_uscite',
                              'pivot_giugno_uscite',
                              'pivot_luglio_uscite',
                              'pivot_agosto_uscite',
                              'pivot_settembre_uscite',
                              'pivot_ottobre_uscite',
                              'pivot_novembre_uscite',
                              'pivot_dicembre_uscite'
                              ]

# pivot_table_w_subtotals
#       (df=list_df_conti_camerino_mese_entrate[0],values='Euro',indices=['Entrate_Uscite', 'Categoria', 'Voce'],columns=[],aggfunc='sum',fill_value='')


i=0
for x in range(12):
    list_pivot_mese_entrate[i] = np.round(pivot_table_w_subtotals
                                 (df=list_df_conti_mese_entrate[i],
                                  values='Euro',
                                  indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                  columns=[],
                                  aggfunc='sum',
                                  fill_value=''), 2)


    list_pivot_mese_uscite[i] = np.round(pivot_table_w_subtotals
                                 (list_df_conti_mese_uscite[i],
                                  values='Euro',
                                  indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                  columns=[],
                                  aggfunc='sum',
                                  fill_value=''), 2)
    i +=1

list_pivot_mese_entrate=[list_pivot_mese_entrate[0], #pivot gennaio entrate
                    list_pivot_mese_entrate[1],
                    list_pivot_mese_entrate[2],
                    list_pivot_mese_entrate[3],
                    list_pivot_mese_entrate[4],
                    list_pivot_mese_entrate[5],
                    list_pivot_mese_entrate[6],
                    list_pivot_mese_entrate[7],
                    list_pivot_mese_entrate[8],
                    list_pivot_mese_entrate[9],
                    list_pivot_mese_entrate[10],
                    list_pivot_mese_entrate[11]
                    ]

list_pivot_mese_uscite= [list_pivot_mese_uscite[0],
                    list_pivot_mese_uscite[1],
                    list_pivot_mese_uscite[2],
                    list_pivot_mese_uscite[3],
                    list_pivot_mese_uscite[4],
                    list_pivot_mese_uscite[5],
                    list_pivot_mese_uscite[6],
                    list_pivot_mese_uscite[7],
                    list_pivot_mese_uscite[8],
                    list_pivot_mese_uscite[9],
                    list_pivot_mese_uscite[10],
                    list_pivot_mese_uscite[11]
                    ]

#print(list_pivot_mese_entrate[0])

##############PRIMA PAGINA#############################

# Creo il file 'conti_styled.xlsx'
wb = Workbook()
# La prima pagina 'Sheet' la chiamo 'Copertina_fronte'
wb['Sheet'].title = ('Copertina_fronte')

wb['Copertina_fronte'].merge_cells('A4:I4')
wb['Copertina_fronte']['A4'] = 'Resoconto Amministrativo'
wb['Copertina_fronte']['A4'].font = Font(name='Calibri',
                                size=35,
                                bold=True,
                                italic=True,
                                vertAlign='none',
                                underline='single',
                                strike=False,
                                color='204ac8') #blu royal
wb['Copertina_fronte']['A4'].alignment = Alignment(horizontal='center')

wb['Copertina_fronte']['A7'] = 'Fraternità di .....'
wb['Copertina_fronte']['A7'].font = Font(name='Calibri',
                                size=25,
                                bold=True,
                                italic=True,
                                vertAlign='none',
                                underline='single',
                                strike=False,
                                color='204ac8')

wb['Copertina_fronte']['A10'] = 'Anno.....'
wb['Copertina_fronte']['A10'].font = Font(name='Calibri',
                                size=25,
                                bold=True,
                                italic=True,
                                vertAlign='none',
                                underline='single',
                                strike=False,
                                color='204ac8')



# Inserisco immagine bilancia
img = openpyxl.drawing.image.Image('bilancia.png')
img.anchor = 'B13'
wb['Copertina_fronte'].add_image(img)

wb.save('conti_styled.xlsx')
##################################

# Con ExcelWriter di pandas METTO INSIEME il pivot delle entrate e il pivot delle uscite


# sheets dei 12 mesi
list_ws_mese = ['ws_gennaio',
                'ws_febbraio',
                'ws_marzo',
                'ws_aprile',
                'ws_maggio',
                'ws_giugno',
                'ws_luglio',
                'ws_agosto',
                'ws_settembre',
                'ws_ottobre',
                'ws_novembre',
                'ws_dicembre'
                ]

c = 0 # contatore

for x in range (12):
    with pd.ExcelWriter('conti_styled.xlsx',
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                        ) as writer:
        list_pivot_mese_entrate[c].to_excel(writer, sheet_name=list_mese[c], startrow=5)
        list_pivot_mese_uscite[c].to_excel(writer, sheet_name=list_mese[c], startrow=(len(list_pivot_mese_entrate[c] ) + 10))

    # leggo il file "conti_camerino_styled.xlsx"
    wb = load_workbook(filename="conti_styled.xlsx")
    #creo i 12 sheet
    list_ws_mese[c]= wb[list_mese[c]]

    c += 1



#sheets dei 12 mesi
list_ws_mese = [wb[list_mese[0]],  #ws_gennaio,
                wb[list_mese[1]],  #ws_febbraio,
                wb[list_mese[2]],  #ws_marzo,
                wb[list_mese[3]],  #ws_aprile,
                wb[list_mese[4]],  #ws_maggio,
                wb[list_mese[5]],  #ws_giugno,
                wb[list_mese[6]],  #ws_luglio,
                wb[list_mese[7]],  #ws_agosto,
                wb[list_mese[8]],  #ws_settembre,
                wb[list_mese[9]],  #ws_ottobre,
                wb[list_mese[10]],  #ws_novembre,
                wb[list_mese[11]],  #ws_dicembre
                ]

################# APPLICO STILE ########################



# Colonna D :Formattazione degli euro in valuta euro
for sheet in list_ws_mese:
    for row in sheet[7:sheet.max_row]:  # skip the header
        #print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
        cell = row[3]  # il quarto valore della tuple
        print (cell)# <Cell 'multiple'.D7>
        cell.number_format = '#,##0.00 €'
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)

# Aggiungo la scritta 'Totale =' alla tabella pivot davanti ai subtotali
for sheet in list_ws_mese:
    for row in sheet[7:sheet.max_row]:
        for cell in sheet['B']: #per ogni casella della colonna B
            if cell.value is not None:
                # ossia se la casella nella colonna B non è vuota
                # Assegnale uno stile
                cell.font = Font(name='Calibri',
                                size=15,
                                bold=True,
                                italic=True,
                                vertAlign='none',
                                underline='single',
                                strike=False,
                                color='a81a1a')
                # Assegna uno stile anche alla cell accanto corrispondente
                sheet.cell(row=cell.offset(row=0, column=0).row, column=3,
                                    value=f"Totale {cell.value} =")
                sheet.cell(row=cell.offset(row=0, column=0).row, column=4).font\
                                    = Font(size=15, color='a81a1a', bold=True)
                sheet.cell(row=cell.offset(row=0, column=0).row, column=4).number_format\
                                    = '#,##0.00 €'
                sheet.cell(row=cell.offset(row=0, column=0).row, column=4).alignment\
                                    = Alignment(horizontal="left")



# Larghezza fissa colonne
i = 0  # contatore
# set the height of the first row in each sheet
for sheet in list_ws_mese:
    sheet.row_dimensions[1].height = 70

    # set the width of the column
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15

    # merge cells
    sheet.merge_cells('A1:D1')

    # scrivo nella cella 'A1'
    sheet['A1'].value = list_mese[i]
    i += 1

    # Formattazione cella
    sheet['A1'].font = Font(name='Calibri',
                            size=25,
                            bold=True,
                            italic=True,
                            vertAlign='none',
                            underline='single',
                            strike=False,
                            color='a81a1a')

    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")



    # Colonna C: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[2]  # il terzo valore della tuple
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Colonna D: Allineamento
    for row in sheet[7:sheet.max_row]:  # skip the header
        cell = row[1]  # il secondo valore della tuple
        cell.alignment = Alignment(horizontal="center", vertical="center")



    # Formattazione headers
    list = []

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("Categoria") or
                cell.value == ("Entrate") or
                cell.value == ("Euro") or
                cell.value == ("Uscite") or
                cell.value == ("Voce")):
                list.append(cell)
    for cell in list:
        cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.rows:
        for cell in row:
            if (cell.value == ("Totale Categoria =")):

                    cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")


    # Formattazione ' Euro accanto a 'Totale categoria='

    double = Side(border_style="double", color="4617F1")
    cont = 0
    for row in sheet.rows:
        for cell in row:
            if (cell.value == ('Totale Categoria =')):
                    #
                    sheet.cell(row=cell.offset(row=0, column=1).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                    sheet.cell(row=cell.offset(row=0, column=1).row, column=4).number_format \
                            = '#,##0.00 €'
                    sheet.cell(row=cell.offset(row=0, column=1).row, column=4).alignment \
                            = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=0, column=1).row, column=4).fill\
                        = PatternFill('solid', fgColor='d1d22e')
                    sheet.cell(row=cell.offset(row=0, column=1).row, column=4).border\
                        = Border(bottom=double, top=double, left=double, right=double)

                    sheet.cell(row=cell.offset(row=1, column=1).row, column=4).font \
                        = Font(size=15, color='a81a1a', bold=True)
                    sheet.cell(row=cell.offset(row=1, column=1).row, column=4).number_format \
                        = '#,##0.00 €'
                    sheet.cell(row=cell.offset(row=1, column=1).row, column=4).alignment \
                        = Alignment(horizontal="left")
                    sheet.cell(row=cell.offset(row=1, column=1).row, column=4).fill \
                        = PatternFill('solid', fgColor='d1d22e')
                    sheet.cell(row=cell.offset(row=1, column=1).row, column=4).border \
                        = Border(bottom=double, top=double, left=double, right=double)

    # Rendi 'invisibile il testo"Entrate_Uscite"
    list = []

    # Formattazione "Entrate_Uscite", piccolo per non essere visto

    for row in sheet.rows:
        for cell in row:
            if cell.value == ("Entrate_Uscite"):
                list.append(cell)
    for cell in list:
        cell.font = Font(size=1)






wb.save('conti_styled.xlsx')