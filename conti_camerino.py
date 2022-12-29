# provo ad importare conti_camerino
import pandas as pd
#leggo e creo dataframe sensa indici colonna
df_conti_camerino=pd.read_excel('conti_camerino_da_importare.xlsx', header=None, index_col=None)
#verifico
#print(df_conti_camerino.head())
print() # riga vuota

#imposto gli indici colonna
df_conti_camerino.columns = ['Anno', 'Mese', 'Categoria', 'Voce','Euro']
#verifico
#print(df_conti_camerino.head())
print()

#salvo il file con un nome diverso
df_conti_camerino.to_excel('conti_camerino_modified.xlsx')
# Leggo il nuovo file, creo nuovo dataframe senza colonna indice
df_conti_camerino_modified = pd.read_excel('conti_camerino_modified.xlsx', index_col=0)

# Stampo le prime 5 righe. Potrei anche usare .head()
#print(df_conti_camerino_modified.iloc[0:5])

#print(df_conti_camerino_modified.head())
#print(df_conti_camerino_modified.info())

#seleziona solo le colonne desiderate
#wanted_columns = df_conti_camerino_modified[['Anno', 'Mese']]
#salva il fle come hai fatto prima con un nome diverso .to_excel
# Leggo il nuovo file, creo nuovo dataframe senza colonna indice



#pulizia delle righe con valori nulli
# print(df_conti_camerino_modified.isnull().sum())
# print(df_conti_camerino_modified.dropna(inplace=True))
#print(df_conti_camerino_modified.isnull().sum())



# print(df_conti_camerino_modified.loc
#       [df_conti_camerino_modified['Categoria']=='Sussidi'])



df_conti_camerino_modified['Entrate_Uscite']= df_conti_camerino_modified['Categoria'].apply\
    (lambda x: 'Entrate'    if  x=='Vendite varie' or
                                x=='Salute' or
                                x=='Curia' or
                                x=='Collette-Chiesa' or
                                x=='Congrua' or
                                x=='Interessi' or
                                x=='Messe celebrate' or
                                x=='Offerte' or
                                x=='Pensioni' or
                                x=='Predicazione' or
                                x=='Servizi religiosi' or
                                x=='Stipendi' or
                                x=='Sussidi' or
                                x=='Rimbosi' or
                                x=='Vendite varie' or
                                x=='Eccedenza Cassa'
                            else 'Uscite')


df_conti_camerino_modified = df_conti_camerino_modified [['Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce','Euro']]
#print(df_conti_camerino_modified.head())

#df_conti_camerino_modified.to_excel('conti_camerino_modified_excel.xlsx')

########PIVOT
import numpy as np
#df_conti_camerino_modified['Anno'] = df_conti_camerino_modified['Anno'].astype(str)
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].astype("category")
df_conti_camerino_modified["Mese"] = df_conti_camerino_modified["Mese"].cat.set_categories(["gennaio","febbraio","marzo","aprile","maggio","giugno","luglio","agosto","settembre","ottobre","novembre","dicembre"])

# Example 1 - Using loc[] with multiple conditions
                                                                    # df2=df.loc[(df['Discount'] >= 1000) & (df['Discount'] <= 2000)]
#
# # Example 2
# df2=df.loc[(df['Discount'] >= 1200) | (df['Fee'] >= 23000 )]
# print(df2)







df_conti_camerino_pivot_tabellone_anno_entrate = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'gennaio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Entrate')]

df_conti_camerino_pivot_tabellone_anno_uscite = df_conti_camerino_modified.loc[
                                        (df_conti_camerino_modified['Anno'] == 2015) &
                                        (df_conti_camerino_modified['Mese'] == 'gennaio') &
                                        (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]




print(df_conti_camerino_pivot_tabellone_anno_entrate.head())

pivot_gennaio_entrate = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_tabellone_anno_entrate,
                               values='Euro',
                               index=['Entrate_Uscite', 'Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE',
                               fill_value=0),2)




pivot_gennaio_uscite = np.round(pd.pivot_table
                            (df_conti_camerino_pivot_tabellone_anno_uscite,
                               values='Euro',
                               index=['Entrate_Uscite','Categoria','Voce'],
                               aggfunc='sum',
                               margins=True,
                               margins_name= 'TOTALE',
                               fill_value=0),2)

print(pivot_gennaio_entrate)
print(pivot_gennaio_uscite)


# Create a Pandas Excel writer using XlsxWriter as the engine.
#writer = pd.ExcelWriter('conti_camerino_multiple.xlsx')

# oppure

with pd.ExcelWriter("conti_camerino_multiple.xlsx") as writer:
    pivot_gennaio_uscite.to_excel(writer, sheet_name='gennaio_uscite')
    pivot_gennaio_entrate.to_excel(writer, sheet_name='gennaio_entrate')

# with pd.ExcelWriter("conti_camerino_multiple.xlsx", mode="a", engine="openpyxl") as writer:
#     df_conti_camerino_pivot_tabellone_anno_entrate.to_excel(writer, sheet_name="mode a")  # doctest: +SKIP

with pd.ExcelWriter("conti_camerino_multiple.xlsx",
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                    ) as writer:
                    pivot_gennaio_entrate.to_excel(writer, sheet_name="multiple", startrow=5)
                    pivot_gennaio_uscite.to_excel(writer, sheet_name="multiple", startrow=(len(pivot_gennaio_entrate)+10))

########################
########################
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
wb = load_workbook(filename = "conti_camerino_multiple.xlsx")
ws_entrate = wb['gennaio_entrate']
ws_uscite=wb['gennaio_uscite']
ws_multiple=wb['multiple']

# set the height of the row
ws_multiple.row_dimensions[1].height = 70

# set the width of the column
ws_multiple.column_dimensions['A'].width = 15
ws_multiple.column_dimensions['B'].width = 20
ws_multiple.column_dimensions['C'].width = 30
ws_multiple.column_dimensions['D'].width = 15


from openpyxl.styles import numbers

#ws_multiple['D21'].number_format = '_-* #.##0,00 €_-;-* #.##0,00 €_-;_-* "-"?? €_-;_-@_-'
# ws_multiple['D21'].number_format = '_-* #.##0,00 €_-;-* #.##0,00 €_-;_-* "-"?? €_-;_-@_-'
# ws_multiple.number_format = numbers.FORMAT_PERCENTAGE
center_align = Alignment(horizontal='center', vertical='center')
###################




# for c in ws_multiple['D']:
#     # c[0].border = no_right_side
#     # c[0].alignment = center_align
#     ws_multiple.number_format = '_-* #.##0,00 €_-;-* #.##0,00 €_-;_-* "-"?? €_-;_-@_-'

ws_multiple.merge_cells('A1:D1')
top_left_cell = ws_multiple['A1']
top_left_cell.value = "Conti mese di gennaio"
top_left_cell.font=Font(name='Calibri',
                        size=25,
                        bold=True,
                        italic=True,
                        vertAlign='none',
                        underline='single',
                        strike=False,
                        color='a81a1a')


top_left_cell.alignment = Alignment(horizontal="center", vertical="center")


print('type')
print(wb.sheetnames)
print(type(ws_uscite))
print(type(ws_entrate))
print(type(ws_multiple))

print('type')

print('valoreeee')
print(ws_entrate['D18'].value)
print(ws_uscite['D18'].value)
print(ws_multiple['D8'].value)

# center align column H in the default sheet:
#ws = wb.active

for row in ws_multiple[7:ws_multiple.max_row]:  # skip the header
    #print(row) #(<Cell 'multiple'.A7>, <Cell 'multiple'.B7>, <Cell 'multiple'.C7>, <Cell 'multiple'.D7>)
    cell = row[3] #il quarto valore della tuple
    #print (cell)# <Cell 'multiple'.D7>
    cell.number_format = '#,##0.00€'
    cell.alignment = Alignment(horizontal="right")
    cell.font=Font(bold=True)

for row in ws_multiple[7:ws_multiple.max_row]:  # skip the header
    cell = row[2]  #il terzo valore della tuple
    cell.alignment = Alignment(horizontal="right")

for row in ws_multiple[7:ws_multiple.max_row]:  # skip the header
    cell = row[1]  #il secondo valore della tuple
    cell.alignment = Alignment(horizontal="center", vertical="center")


list=[]
# Enumerate the cells in the second row
for row in ws_multiple.rows:
    for cell in row:
        if (cell.value == ("Categoria") or
            cell.value == ("Entrate") or
            cell.value == ("Euro") or
            cell.value == ("Uscite") or
            cell.value == ('TOTALE') or
            cell.value == ("Voce")):
            print('trovato')
            print(cell)
            list.append(cell)
print(list)

for cell in list:
    cell.font = Font(size=15, color='a81a1a', bold=True)



list=[]
# Enumerate the cells in the second row
for row in ws_multiple.rows:
    for cell in row:
        if cell.value == ("Entrate_Uscite"):
            list.append(cell)
for cell in list:
    cell.font = Font(size=1)

# for cell in list:
#     cell.font = Font(size=1)
#     print(cell)
#     print(cell.coordinate)
#     print(cell.row)
#     print(cell.column)

list=[]

for row in ws_multiple.rows:
    for cell in row:
        if cell.value == ("TOTALE"):
            list.append(cell)
for cell in list:
    ws_multiple.cell(cell.row, column=4).font = Font(size=15, color='a81a1a', bold=True)


wb.save("conti_camerino_styled.xlsx")