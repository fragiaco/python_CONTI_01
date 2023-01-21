from tkinter import *
from tkinter import ttk
from tkinter import messagebox

import sqlite3
import pandas as pd
import os, sys, subprocess

from openpyxl.styles import Font, Alignment
from openpyxl.styles import Side, Border

from openpyxl import styles, formatting
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# Leggi il file xlsx e trasformalo in dataframe impostando i nomi colonna
from openpyxl.worksheet.worksheet import Worksheet

from xlsxwriter.utility import xl_rowcol_to_cell


######################################################################
######## FUNZIONE CONNESSIONE AL DATABASE 'database_conti' ###########
######################################################################
def connessione():
    conn = sqlite3.connect('database_conti')

    cur = conn.cursor()
    try:
        cur.execute('''CREATE TABLE TABLE_Conti(ID integer not null PRIMARY KEY ,
                                                Anno TEXT not null ,
                                                Mese TEXT not null ,
                                                Entrate_Uscite TEXT not null ,
                                                Categoria TEXT not null ,
                                                Voce TEXT not null ,
                                                Euro real not null )''')
    except:
        pass

    print(conn)
    print('Sei connesso al database_conti')
    conn.commit()
    conn.close()


connessione()

############################
######## TKINTER ###########
############################

root = Tk()

# DEFINISCO le dimensioni della GUI e il TITOLO nella barra superiore
height = 950  # altezza
width = 1680  # larghezza
top = 0
left = (root.winfo_screenwidth() - width) / 2
geometry = ("{}x{}+{}+{}".format(width, height, int(left), int(top)))
root.geometry(geometry)
root.resizable(0, 0)
root.title('Conti Convento Camerino')

# Label title
title = Label(root, text='Database Entrate Uscite', font=('verdana', 40, 'bold'), bg='blue', fg='#ffeddb')
title.pack(side=TOP, fill=X)

###################################
######## TKINTER frames ###########
###################################
# LATO SINISTRO
# Frame 1 - left side Frame
Frame1 = Frame(root, bd='4', bg='blue', relief=RIDGE)
Frame1.place(x=20, y=85, width=550, height=850)
# Frame 1in - calcolatrice
Frame_calc = Frame(Frame1, bd='4', bg='light blue', relief=RIDGE)
Frame_calc.grid(column=0, row=15, columnspan=2, padx=65, pady=70)
# Frame 1in - bottom side Frame - Button 'add'
Frame1in = Frame(Frame1, bd='4', bg='blue', relief=RIDGE)
Frame1in.place(x=15, y=768, width=500, height=60)

# LATO DESTRO
# Frame 2 - right side Frame
Frame2 = Frame(root, bd='4', bg='blue', relief=RIDGE)
Frame2.place(x=590, y=85, width=1070, height=850)
# Frame 2in - treeview right Frame
Frame2in_tree = Frame(Frame2, bd='4', bg='blue', relief=RIDGE)
Frame2in_tree.place(x=15, y=15, width=1015, height=335)
# Frame 2in - Place holder COMBOBOX e EXCELL FRAMES e BUTTONS
Frame2in_bottom = Frame(Frame2, bd='4', bg='blue', relief=RIDGE)
Frame2in_bottom.place(x=15, y=360, width=1015, height=480)
# Frame 2in - combobox
Frame_combobox_ok = Frame(Frame2in_bottom, bd='4', bg='blue', relief=RIDGE)
Frame_combobox_ok.place(x=15, y=2, width=485, height=400)
# Frame 2in - excell
Frame_excell = Frame(Frame2in_bottom, bd='4', bg='blue', relief=RIDGE)
Frame_excell.place(x=500, y=2, width=485, height=400)
# Frame 2in -  update botton
Frame_update_botton = Frame(Frame2in_bottom, bd='4', bg='blue', relief=RIDGE)
Frame_update_botton.place(x=15, y=405, width=485, height=60)
# Frame 2in -  excell botton
Frame_excell_botton = Frame(Frame2in_bottom, bd='4', bg='blue', relief=RIDGE)
Frame_excell_botton.place(x=500, y=405, width=485, height=60)
#######################################################


# define mylabel
mylabel = Label(Frame2)

##########################
# Frame1 Labels
Frame1_title = Label(Frame1, text='Inserisci Dati:', font=('verdana', 20, 'bold'), bg='blue', fg='#ffff66')
Frame1_title.grid(row=0, columnspan=2, padx=20, pady=10, sticky='w')

Frame1_anno = Label(Frame1, text='Anno', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_anno.grid(row=2, padx=20, pady=10, sticky='w')

Frame1_mese = Label(Frame1, text='Mese', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_mese.grid(row=4, padx=20, pady=25, sticky='w')

Frame1_EntrateUscite = Label(Frame1, text='Entrate/Uscite', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_EntrateUscite.grid(row=6, padx=25, pady=20, sticky='w')

Frame1_Categoria = Label(Frame1, text='Categoria', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_Categoria.grid(row=8, padx=25, pady=20, sticky='w')

Frame1_VoceSpesa = Label(Frame1, text='Voce di spesa', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_VoceSpesa.grid(row=10, padx=25, pady=20, sticky='w')

Frame1_Euro = Label(Frame1, text='Euro', font=('verdana', 15, 'bold'), bg='blue', fg='white')
Frame1_Euro.grid(row=12, padx=25, pady=20, sticky='w')

######## CALCOLATRICE #########

# Python program to create a simple GUI
# calculator using Tkinter

# globally declare the expression variable
expression = ""


# Function to update expression
# in the text entry box
def press(num):
    # point out the global expression variable
    global expression

    # concatenation of string
    expression = expression + str(num)

    # update the expression by using set method
    equation.set(expression)


# Function to evaluate the final expression
def equalpress():
    # Try and except statement is used
    # for handling the errors like zero
    # division error etc.

    # Put that code inside the try block
    # which may generate the error
    try:

        global expression

        # eval function evaluate the expression
        # and str function convert the result
        # into string
        total = str(eval(expression))

        equation.set(total)

        # initialize the expression variable
        # by empty string
        expression = ""

    # if error is generate then handle
    # by the except block
    except:

        equation.set(" error ")
        expression = ""


# Function to clear the contents
# of text entry box
def clear():
    global expression
    expression = ""
    equation.set("")


# StringVar() is the variable class
# we create an instance of this class
equation = StringVar()

# create the text entry box for
# showing the expression .
expression_field = Entry(Frame_calc, textvariable=equation)

# grid method is used for placing
# the widgets at respective positions
# in table like structure .
# expression_field.grid(columnspan=4, ipadx=70)

# create a Buttons and place at a particular
# location inside the root window .
# when user press the button, the command or
# function affiliated to that button is executed .
button1 = Button(Frame_calc, text=' 1 ', fg='black', bg='light grey',
                 command=lambda: press(1), height=1, width=7)
button1.grid(row=2, column=0, pady=2)

button2 = Button(Frame_calc, text=' 2 ', fg='black', bg='light grey',
                 command=lambda: press(2), height=1, width=7)
button2.grid(row=2, column=1)

button3 = Button(Frame_calc, text=' 3 ', fg='black', bg='light grey',
                 command=lambda: press(3), height=1, width=7)
button3.grid(row=2, column=2)

button4 = Button(Frame_calc, text=' 4 ', fg='black', bg='light grey',
                 command=lambda: press(4), height=1, width=7)
button4.grid(row=3, column=0, pady=2)

button5 = Button(Frame_calc, text=' 5 ', fg='black', bg='light grey',
                 command=lambda: press(5), height=1, width=7)
button5.grid(row=3, column=1)

button6 = Button(Frame_calc, text=' 6 ', fg='black', bg='light grey',
                 command=lambda: press(6), height=1, width=7)
button6.grid(row=3, column=2)

button7 = Button(Frame_calc, text=' 7 ', fg='black', bg='light grey',
                 command=lambda: press(7), height=1, width=7)
button7.grid(row=4, column=0, pady=2)

button8 = Button(Frame_calc, text=' 8 ', fg='black', bg='light grey',
                 command=lambda: press(8), height=1, width=7)
button8.grid(row=4, column=1)

button9 = Button(Frame_calc, text=' 9 ', fg='black', bg='light grey',
                 command=lambda: press(9), height=1, width=7)
button9.grid(row=4, column=2)

button0 = Button(Frame_calc, text=' 0 ', fg='black', bg='light grey',
                 command=lambda: press(0), height=1, width=7)
button0.grid(row=5, column=0, pady=2)

plus = Button(Frame_calc, text=' + ', fg='black', bg='light grey',
              command=lambda: press("+"), height=1, width=7)
plus.grid(row=2, column=3)

minus = Button(Frame_calc, text=' - ', fg='black', bg='light grey',
               command=lambda: press("-"), height=1, width=7)
minus.grid(row=3, column=3)

multiply = Button(Frame_calc, text=' * ', fg='black', bg='light grey',
                  command=lambda: press("*"), height=1, width=7)
multiply.grid(row=4, column=3)

divide = Button(Frame_calc, text=' / ', fg='black', bg='light grey',
                command=lambda: press("/"), height=1, width=7)
divide.grid(row=5, column=3)

equal = Button(Frame_calc, text=' = ', fg='black', bg='light grey',
               command=equalpress, height=1, width=7)
equal.grid(row=5, column=2)

clear = Button(Frame_calc, text='Clear', fg='black', bg='light grey',
               command=clear, height=1, width=7)
clear.grid(row=5, column=1)

Decimal = Button(Frame_calc, text='.', fg='black', bg='light grey',
                 command=lambda: press('.'), height=1, width=7)
Decimal.grid(row=6, column=0)
# start the GUI
# gui.mainloop()

############################
####### COMBOBOX ###########
############################

# List Anno
Anni = ["2022",
        "2023",
        "2024",
        "2025",
        ]

# List Mesi
Mesi = ["gennaio",
        "febbraio",
        "marzo",
        "aprile",
        "maggio",
        "giugno",
        "luglio",
        "agosto",
        "settembre",
        "ottobre",
        "novembre",
        "dicembre",
        ]

# List Entrate_Uscite
Entrate_Uscite = ["Entrate",
                  "Uscite",
                  ]

# List Categorie_Entrate
Categorie_Entrate = ["Collette_Chiesa",
                     "Congrua",
                     "Interessi",
                     "Messe_celebrate",
                     "Offerte",
                     "Pensioni",
                     "Servizi_religiosi",
                     ]

# List Voci Entrate
Collette_Chiesa = ["Cestino",
                   "Cassette"
                   ]
Congrua = ['fra Giacomo'
           ]
Interessi = ['Interessi bancari'
             ]
Messe_celebrate = ['Messe_celebrate'
                   ]
Offerte = ['Eccedenza_Messe',
           'Offerte libere'
           ]
Pensioni = ['fra Gabriele'
            ]
Servizi_religiosi = ['Servizi_religiosi',
                     'Predicazione'
                     ]

# List Categorie Uscite
Categorie_Uscite = ["Acquisti_Chiesa",
                    "Acquisti_Convento",
                    "Acquisti_Orto_Animali",
                    "Cultura",
                    "Curia_provinciale",
                    "Domestici",
                    "Elargizioni",
                    "Utenze",
                    "Ferie_Viaggi",
                    "Igiene",
                    "Imposte",
                    "Lavori_Impianti",
                    "Posta_Cancelleria",
                    "Salute",
                    "Veicoli_motore",
                    "Vestiario",
                    'Vitto',
                    'Eccedenza_Cassa',
                    ]
# List Voci Entrate
Acquisti_Chiesa = ['Fiori',
                   'Ostie',
                   'Ceri e Candele',
                   'Paramenti liturigici'
                   ]
Acquisti_Convento = ['Ferramenta',
                     'Materiale elettrico',
                     'Casalinghi',
                     'Materiale edile',
                     'Computer',
                     'Stampante'
                     ]
Acquisti_Orto_Animali = ['Attrezzi agricoli manutenzione',
                         'Semi, Ortaggi',
                         'Fitofarmaci'
                         ]
Cultura = ['Abbonamenti',
           'Convegni',
           'Ritiro spirituale',
           'Libri_Riviste',
           'Pellegrinaggio'
           ]
Curia_provinciale = ['Curia_provinciale',
                     'Tassa_Curia_generale'
                     ]
Domestici = ['Rosaria'
             ]
Elargizioni = ['Regalie',
               'Beneficenza_5%'
               ]
Utenze = ['Energia_elettrica',
          'Gas',
          'Acqua',
          'Riscaldamento',
          'Rifiuti',
          'pay_TV',
          'Telefono_Internet'
          ]
Ferie_Viaggi = ['Carburante',
                'Treno',
                'Aereo',
                'Autostrada-Parcheggio'
                ]
Igiene = ['Detersivi',
          'Igiene personale',
          ]
Imposte = ['Imposte_bancarie',
           'Controllo_estintori'
           'Imposte_varie'
           ]
Lavori_Impianti = ['Elettricista',
                   'Fabbro',
                   'Idraulico',
                   'Muratore',
                   'Imbianchino',
                   'Restauratore'
                   ]
Posta_Cancelleria = ['Cancelleria',
                     'Posta'
                     ]
Salute = ['Ticket_Esami',
          'Farmacia'
          ]
Veicoli_motore = ['Assicurazione',
                  'Bollo Auto',
                  'Meccanico',
                  'Gommista',
                  'Elettrauto',
                  'Carrozziere'
                  ]
Vestiario = ['Indumenti'
             ]
Vitto = ['Alimentari'
         ]
Eccedenza_Cassa = ['Eccedenza_Cassa'
                   ]


def pick_Categoria(e):
    if my_combo.get() == "Entrate":
        categoria_combo.config(values=Categorie_Entrate)
        categoria_combo.set('click!')
        categoria_combo['state'] = 'readonly'

    if my_combo.get() == "Uscite":
        categoria_combo.config(values=Categorie_Uscite)
        categoria_combo.set('click!')
        categoria_combo['state'] = 'readonly'


def pick_Voce(e):
    # VOCI ENTRATA
    if categoria_combo.get() == "Collette_Chiesa":
        voce_combo.config(values=Collette_Chiesa)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Congrua":
        voce_combo.config(values=Congrua)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Interessi":
        voce_combo.config(values=Interessi)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Messe celebrate":
        voce_combo.config(values=Messe_celebrate)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Offerte":
        voce_combo.config(values=Offerte)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Pensioni":
        voce_combo.config(values=Pensioni)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Servizi_religiosi":
        voce_combo.config(values=Servizi_religiosi)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    # VOCI USCITA

    if categoria_combo.get() == "Acquisti_Chiesa":
        voce_combo.config(values=Acquisti_Chiesa)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Acquisti_Convento":
        voce_combo.config(values=Acquisti_Convento)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Acquisti_Orto_Animali":
        voce_combo.config(values=Acquisti_Orto_Animali)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Cultura":
        voce_combo.config(values=Cultura)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Curia_provinciale":
        voce_combo.config(values=Curia_provinciale)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Domestici":
        voce_combo.config(values=Domestici)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Elargizioni":
        voce_combo.config(values=Elargizioni)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Utenze":
        voce_combo.config(values=Utenze)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Ferie_Viaggi":
        voce_combo.config(values=Ferie_Viaggi)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Igiene":
        voce_combo.config(values=Igiene)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Imposte":
        voce_combo.config(values=Imposte)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Lavori_Impianti":
        voce_combo.config(values=Lavori_Impianti)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Posta_Cancelleria":
        voce_combo.config(values=Posta_Cancelleria)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Salute":
        voce_combo.config(values=Salute)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Veicoli_motore":
        voce_combo.config(values=Veicoli_motore)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Vestiario":
        voce_combo.config(values=Vestiario)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Vitto":
        voce_combo.config(values=Vitto)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'

    if categoria_combo.get() == "Eccedenza_Cassa":
        voce_combo.config(values=Eccedenza_Cassa)
        voce_combo.set('click!')
        voce_combo['state'] = 'readonly'


# Dropbox Anno
anno_combo = ttk.Combobox(Frame1, font=("Helvetica", 15), values=Anni)
anno_combo.current(0)
anno_combo.grid(row=2, column=1)
anno_combo['state'] = 'readonly'

# Dropbox Mesi
mesi_combo = ttk.Combobox(Frame1, font=("Helvetica", 15), values=Mesi)
mesi_combo.current(0)
mesi_combo.grid(row=4, column=1)
mesi_combo['state'] = 'readonly'

# Dropbox Entrate_Uscite
my_combo = ttk.Combobox(Frame1, font=("Helvetica", 15), values=Entrate_Uscite)
my_combo.set('click!')
my_combo.grid(row=6, column=1)
my_combo['state'] = 'readonly'

# Bind the ComboBox
my_combo.bind("<<ComboboxSelected>>", pick_Categoria)

# Categoria ComboBox
categoria_combo = ttk.Combobox(Frame1, font=("Helvetica", 15), values=[""])
categoria_combo.current(0)
categoria_combo.grid(row=8, column=1)
categoria_combo['state'] = 'readonly'

# Bind the ComboBox
categoria_combo.bind("<<ComboboxSelected>>", pick_Voce)

# Voce Entrata_Spesa ComboBox Combo Box
voce_combo = ttk.Combobox(Frame1, font=("Helvetica", 15), values=[""])
voce_combo.current(0)
voce_combo.grid(row=10, column=1)
voce_combo['state'] = 'readonly'

# euro ENTRY
# euro = Entry(Frame1, font=("Helvetica", 15, 'bold'), bd=5, relief=GROOVE)
euro = Entry(Frame1, font=("Helvetica", 15, 'bold'), bd=5, relief=GROOVE, textvariable=equation)
euro.grid(row=12, column=1)

############################
####### TREEVIEW ###########
############################

# Add some style
style = ttk.Style()
# Pick a theme
style.theme_use("default")

# Configure our treeview colors
style.configure("Treeview",
                background="#D3D3D3",
                foreground="black",
                rowheight=30,
                fieldbackground="#D3D3D3",
                font=('Calibri', 12)
                )

# Headings
style.configure("Treeview.Heading",
                font=('Calibri', 12, 'bold')
                )

# Change selected color
style.map('Treeview',
          background=[('selected', 'blue')]
          )

# Treeview Scrollbar
tree_scroll = Scrollbar(Frame2in_tree)
tree_scroll.pack(side=RIGHT, fill=Y)

# Create Treeview
my_tree = ttk.Treeview(Frame2in_tree, yscrollcommand=tree_scroll.set, selectmode="extended")
# Pack to the screen
my_tree.pack()

# Configure the scrollbar
tree_scroll.config(command=my_tree.yview)

# Define Our Columns
my_tree['columns'] = ("Id", "Anno", "Mese", "Entrate_Uscite", "Categoria", "Voce", "Euro")

# Formate Our Columns
my_tree.column("#0", width=0, stretch=NO)
my_tree.column("Id", anchor=CENTER, width=50)
my_tree.column("Anno", anchor=CENTER, width=80)
my_tree.column("Mese", anchor=CENTER, width=80)
my_tree.column("Entrate_Uscite", anchor=CENTER, width=160)
my_tree.column("Categoria", anchor=W, width=200)
my_tree.column("Voce", anchor=W, width=250)
my_tree.column("Euro", anchor=W, width=160)

# Create Headings
my_tree.heading("#0", text="", anchor=W)
my_tree.heading("Id", text="Id", anchor=CENTER)
my_tree.heading("Anno", text="Anno", anchor=CENTER)
my_tree.heading("Mese", text="Mese", anchor=CENTER)
my_tree.heading("Entrate_Uscite", text="Entrate_Uscite", anchor=CENTER)
my_tree.heading("Categoria", text="Categoria", anchor=W)
my_tree.heading("Voce", text="Voce", anchor=W)
my_tree.heading("Euro", text="Euro", anchor=W)


############################
######## SQLITE3 ###########
############################

# Insert into TABLE_Conti
def submit():
    conn = sqlite3.connect('database_conti')
    cur = conn.cursor()

    dati = [(anno_combo.get(), mesi_combo.get(), my_combo.get(), categoria_combo.get(), voce_combo.get(), euro.get())]

    cur.executemany(
        'INSERT INTO TABLE_Conti (Anno, Mese, Entrate_Uscite, Categoria, Voce, Euro) VALUES(?, ?, ?, ? ,? ,?)', dati)
    conn.commit()


# query_database ed insert rows into TREEVIEW
def query_database():
    # Clear the Treeview
    for record in my_tree.get_children():
        my_tree.delete(record)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_conti')

    # Create a cursor instance
    c = conn.cursor()

    c.execute("SELECT * FROM TABLE_Conti")
    records = c.fetchall()

    # for record in records:
    #     print(record)
    # record[0] = id key

    # COLORI RIGHE pari e dispari
    count = 0
    # Create striped row tags
    my_tree.tag_configure('oddrow', background="white")
    my_tree.tag_configure('evenrow', background="lightblue")

    for record in records:
        if count % 2 == 0:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5], record[6]),
                           tags=('evenrow'))
        else:
            my_tree.insert(parent='', index=0, iid=record[0], text='',
                           values=(record[0], record[1], record[2], record[3], record[4], record[5], record[6]),
                           tags=('oddrow'))

        count += 1

    # Al termine del processo la prima riga risulta evidenziata
    child_id = my_tree.get_children()[0]  # la prima riga dall'alto del treeview
    my_tree.focus(child_id)  # evidenziata
    my_tree.selection_set(child_id)

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()


def sqlite3_to_excel():
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_conti')

    # Create a cursor instance
    c = conn.cursor()

    query = "SELECT * FROM TABLE_Conti"  # query to collect recors

    df = pd.read_sql(query, conn)  # create dataframe

    df.sort_values(by='index', ascending=False).to_excel('database_conti.xlsx', index=False, sheet_name='Dati')

    ###########################################################
    ################# Creo il Workbook con OPENPYXL############
    ###########################################################
    wb = Workbook()
    wb = load_workbook(filename="database_conti.xlsx")
    ws = wb.active  # Worksheet

    ws.row_dimensions[1].height = 40
    # openpyxl freeze first row
    ws.freeze_panes = 'A2'
    # openpyxl filter columns
    ws.auto_filter.ref = ws.dimensions

    ############RED################
    red = NamedStyle(name="red")
    red.font = Font(name='Calibri', size=10, color='a81a1a', bold=True)
    red.alignment = Alignment(horizontal="center", vertical="center")
    red.fill = PatternFill('solid', fgColor='d1d22e')
    wb.add_named_style(red)

    ws['A1'].style = 'red'
    ws['B1'].style = 'red'
    ws['C1'].style = 'red'
    ws['D1'].style = 'red'
    ws['E1'].style = 'red'
    ws['F1'].style = 'red'
    ws['G1'].style = 'red'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 21
    ws.column_dimensions['G'].width = 10

    # Colonna G: Formattazione
    for row in ws[2:ws.max_row]:  # skip the header
        cell = row[6]  # il settimo valore della tuple
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0.00 €'

    # ws = wb.create_sheet('Dati')
    wb.save("database_conti_styled.xlsx")

    if sys.platform == "win32":
        os.startfile('database_conti_styled.xlsx')
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, 'database_conti_styled.xlsx'])

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()


###########################################################
#################  EXCEL REPORT  ##########################
###########################################################


### Titolo 'Visualizza dati:'
Frame_excell_title = Label(Frame_excell, text='Visualizza dati:',
                           font=('verdana', 20, 'bold'), bg='blue', fg='#ffff66')
Frame_excell_title.grid(row=0, columnspan=2, padx=10, pady=10, sticky='w')
### Sottotitolo Scrivere prima
Frame_excell_subtitle_anno = Label(Frame_excell, text="Scrivere prima:",
                                   font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=1, columnspan=2, padx=10, pady=10, sticky='w')
### Riga vuota
Frame_excell_subtitle_anno = Label(Frame_excell, text="",
                                   font=('verdana', 1, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=2, columnspan=2, padx=10, pady=10, sticky='w')
### Sottotitolo Anno
Frame_excell_subtitle_anno = Label(Frame_excell, text="Anno di interesse:",
                                   font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno.grid(row=3, columnspan=2, padx=10, pady=10, sticky='w')
### Label Anno
Anno_label_excell = Label(Frame_excell, text="(esempio: 2022)", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Anno_label_excell.grid(row=4, column=0, padx=10, pady=10, sticky='w')
### Entry Anno
Anno_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
Anno_entry_excell.grid(row=4, column=1)
### Sottotitolo Saldo
Frame_excell_subtitle_anno_saldo = Label(Frame_excell, text="Saldo ad inizio anno:",
                                         font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame_excell_subtitle_anno_saldo.grid(row=5, columnspan=2, padx=10, pady=10, sticky='w')
### Label Saldo
Saldo_label_excell = Label(Frame_excell, text="(esempio: 99.99 oppure -99.99)", font=('verdana', 10, 'bold'), bg='blue',
                           fg='white')
Saldo_label_excell.grid(row=6, column=0, padx=10, pady=10, sticky='w')
### Entry Anno
Saldo_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
Saldo_entry_excell.grid(row=6, column=1)

### Sottotitolo Residuo Messe
# Frame_excell_subtitle_residuo_messe = Label(Frame_excell, text="Dis/Avanzo Messe ad inizio anno",
#                             font=('verdana', 10, 'bold'), bg='blue', fg='white')
# Frame_excell_subtitle_residuo_messe.grid(row=7, columnspan=2, padx=10, pady=10, sticky='w')
# ### Label Residuo Messe
# Messe_residuo_label_excell = Label(Frame_excell, text="(esempio: 99 oppure -99)", font=('verdana', 10, 'bold'), bg='blue', fg='white')
# Messe_residuo_label_excell.grid(row=8, column=0, padx=10, pady=10, sticky='w')
# ### Entry Residuo Messe
# Messe_residuo_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable='')
# Messe_residuo_excell.grid(row=8, column=1)

###########################################################
################# COMBOBOX UPDATE #########################
###########################################################

anno_stringvar = StringVar()
mese_stringvar = StringVar()
entrate_uscite_stringvar = StringVar()
categoria_stringvar = StringVar()
voce_stringvar = StringVar()
euro_stringvar = StringVar()

Frame2_bottom_title = Label(Frame_combobox_ok, text='Correggi o Cancella:',
                            font=('verdana', 20, 'bold'), bg='blue', fg='#ffff66')
Frame2_bottom_title.grid(row=0, columnspan=2, padx=10, pady=10, sticky='w')

Frame2_bottom_subtitle = Label(Frame_combobox_ok, text='Selezionare prima una riga nella tabella sopra',
                               font=('verdana', 10, 'bold'), bg='blue', fg='white')
Frame2_bottom_subtitle.grid(row=1, columnspan=2, padx=10, pady=10, sticky='w')

Id_label = Label(Frame_combobox_ok, text="Id", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Id_label.grid(row=2, column=0, padx=10, pady=10, sticky='w')
Id_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', width=17)
Id_entry.grid(row=2, column=1)

Anno_label = Label(Frame_combobox_ok, text="Anno", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Anno_label.grid(row=3, column=0, padx=10, pady=10, sticky='w')
Anno_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=anno_stringvar)
# Anno_entry.grid(row=2, column=1, padx=10, pady=10)

Mese_label = Label(Frame_combobox_ok, text="Mese", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Mese_label.grid(row=4, column=0, padx=10, pady=10, sticky='w')
Mese_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=mese_stringvar)
# Mese_entry.grid(row=3, column=1, padx=10, pady=10)

Entrate_Uscite_label = Label(Frame_combobox_ok, text="Entrate_Uscite", font=('verdana', 10, 'bold'), bg='blue',
                             fg='white')
Entrate_Uscite_label.grid(row=5, column=0, padx=10, pady=10, sticky='w')
Entrate_Uscite_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                             textvariable=entrate_uscite_stringvar)
# Entrate_Uscite_entry.grid(row=4, column=1, padx=10, pady=10)
#
Categoria_label = Label(Frame_combobox_ok, text="Categoria", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Categoria_label.grid(row=6, column=0, padx=10, pady=10, sticky='w')
Categorie_Entrate_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                                textvariable=categoria_stringvar)
# Categorie_Entrate_entry.grid(row=5, column=1, padx=10, pady=10)
#
Voce_label = Label(Frame_combobox_ok, text="Voce", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Voce_label.grid(row=7, column=0, padx=10, pady=10, sticky='w')
Voce_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=voce_stringvar)
# Voce_entry.grid(row=6, column=1, padx=10, pady=10)
#
Euro_label = Label(Frame_combobox_ok, text="Euro", font=('verdana', 10, 'bold'), bg='blue', fg='white')
Euro_label.grid(row=8, column=0, padx=10, pady=10, sticky='w')
Euro_entry = Entry(Frame_combobox_ok, font=('verdana', 10, 'bold'), bg='blue', fg='white', textvariable=euro_stringvar)


# Euro_entry.grid(row=7, column=1, padx=10, pady=10)


def select_record(e):
    # Clear entry boxes
    Id_entry.delete(0, END)
    Anno_entry.delete(0, END)
    Mese_entry.delete(0, END)
    Entrate_Uscite_entry.delete(0, END)
    Categorie_Entrate_entry.delete(0, END)
    Voce_entry.delete(0, END)
    Euro_entry.delete(0, END)

    # Grab record Number
    selected = my_tree.focus()  # focus restituisce l'ID key
    # print(selected) #esempio 38
    # Grab record values
    values = my_tree.item(selected, 'values')
    # print(values) #esempio ('38', '2022', 'gennaio', 'Entrate', 'Messe_celebrate', '', '39.0')

    # outpus to entry boxes
    Id_entry.insert(0, values[0])  # 0 penso significa all'inizio
    Anno_entry.insert(0, values[1])
    Mese_entry.insert(0, values[2])
    Entrate_Uscite_entry.insert(0, values[3])
    Categorie_Entrate_entry.insert(0, values[4])
    Voce_entry.insert(0, values[5])
    Euro_entry.insert(0, values[6])

    # print(Anno_entry.get())


# Bind the treeview
my_tree.bind("<ButtonRelease-1>", select_record)


#######################
def remove_one():
    # x = my_tree.selection()[0] #restituisce l'Id key
    x = my_tree.focus()
    my_tree.delete(x)

    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_conti')

    # Create a cursor instance
    c = conn.cursor()

    # Delete From Database
    c.execute("DELETE from TABLE_Conti WHERE oid=" + Id_entry.get())

    # Commit changes
    conn.commit()

    # Close our connection
    conn.close()

    # Add a little message box for fun
    messagebox.showinfo("Deleted", "Riga Cancellata!")


#######################
######
# Update record
def update_record():
    # Grab the record number
    print('update')
    selected = my_tree.focus()
    print(selected)
    # Update record
    my_tree.item(selected, text="", values=(
    Id_entry.get(), Anno_entry.get(), Mese_entry.get(), Entrate_Uscite_entry.get(), Categorie_Entrate_entry.get(),
    Voce_entry.get(), Euro_entry.get()))

    # Update the database
    # Create a database or connect to one that exists
    conn = sqlite3.connect('database_conti')
    #
    # Create a cursor instance
    c = conn.cursor()

    #
    c.execute("""UPDATE TABLE_Conti SET
    		Anno = :Anno,
    		Mese = :Mese,
    		Entrate_Uscite = :Entrate_Uscite,
    		Categoria = :Categoria,
    		Voce = :Voce,
    		Euro = :Euro

     		WHERE oid = :oid""",
              {
                  'Anno': Anno_entry.get(),
                  'Mese': Mese_entry.get(),
                  'Entrate_Uscite': Entrate_Uscite_entry.get(),
                  'Categoria': Categorie_Entrate_entry.get(),
                  'Voce': Voce_entry.get(),
                  'Euro': Euro_entry.get(),
                  'oid': Id_entry.get()
              })
    #
    #    Commit changes
    conn.commit()
    #
    #         # Close our connection
    conn.close()
    # Add a little message box for fun
    messagebox.showinfo("Updated!", "Riga aggiornata!")

    #         # Clear entry boxes
    Id_entry.delete(0, END)
    Anno_entry.delete(0, END)
    Mese_entry.delete(0, END)
    Entrate_Uscite_entry.delete(0, END)
    Categorie_Entrate_entry.delete(0, END)
    Voce_entry.delete(0, END)
    Euro_entry.delete(0, END)


######
def pick_Categoria_update(e):
    if my_combo_update.get() == "Entrate":
        categoria_combo_update.config(values=Categorie_Entrate)
        # categoria_combo_update.set(entrate_uscite_stringvar.get())
        categoria_combo_update.current(0)

    if my_combo_update.get() == "Uscite":
        categoria_combo_update.config(values=Categorie_Uscite)
        categoria_combo_update.current(0)


def pick_Voce_update(e):
    # VOCI ENTRATA
    if categoria_combo_update.get() == "Collette_Chiesa":
        voce_combo_update.config(values=Collette_Chiesa)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Congrua":
        voce_combo_update.config(values=Congrua)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Interessi":
        voce_combo_update.config(values=Interessi)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Messe celebrate":
        voce_combo_update.config(values=Messe_celebrate)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Offerte":
        voce_combo_update.config(values=Offerte)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Pensioni":
        voce_combo_update.config(values=Pensioni)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Servizi_religiosi":
        voce_combo_update.config(values=Servizi_religiosi)
        voce_combo_update.current(0)

    # VOCI USCITA

    if categoria_combo_update.get() == "Acquisti_Chiesa":
        voce_combo_update.config(values=Acquisti_Chiesa)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Acquisti_Convento":
        voce_combo_update.config(values=Acquisti_Convento)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Acquisti_Orto_Animali":
        voce_combo_update.config(values=Acquisti_Orto_Animali)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Cultura":
        voce_combo_update.config(values=Cultura)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Curia_provinciale":
        voce_combo_update.config(values=Curia_provinciale)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Domestici":
        voce_combo_update.config(values=Domestici)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Elargizioni":
        voce_combo_update.config(values=Elargizioni)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Utenze":
        voce_combo_update.config(values=Utenze)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Ferie_Viaggi":
        voce_combo_update.config(values=Ferie_Viaggi)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == " ":
        voce_combo_update.config(values=Igiene)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Imposte":
        voce_combo_update.config(values=Imposte)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Lavori_Impianti":
        voce_combo_update.config(values=Lavori_Impianti)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Posta_Cancelleria":
        voce_combo_update.config(values=Posta_Cancelleria)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Salute":
        voce_combo_update.config(values=Salute)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Veicoli_motore":
        voce_combo_update.config(values=Veicoli_motore)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Vestiario":
        voce_combo_update.config(values=Vestiario)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Vitto":
        voce_combo_update.config(values=Vitto)
        voce_combo_update.current(0)

    if categoria_combo_update.get() == "Eccedenza_Cassa":
        voce_combo_update.config(values=Eccedenza_Cassa)
        voce_combo_update.current(0)


# Dropbox Anno
anno_combo_update = ttk.Combobox(Frame_combobox_ok, font=("Helvetica", 10), values=Anni, textvariable=anno_stringvar)
anno_combo_update.set(anno_stringvar.get())
# anno_combo_update = ttk.Combobox(Frame2in_bottom, font=("Helvetica", 15), values=Anni, textvariable=new_Anno_entry)
# anno_combo_update.set(new_Anno_entry)
anno_combo_update.grid(row=3, column=1)
# Dropbox Mesi
mesi_combo_update = ttk.Combobox(Frame_combobox_ok, font=("Helvetica", 10), values=Mesi, textvariable=mese_stringvar)
mesi_combo_update.set(mese_stringvar.get())
mesi_combo_update.grid(row=4, column=1)
# Dropbox Entrate_Uscite
my_combo_update = ttk.Combobox(Frame_combobox_ok, font=("Helvetica", 10), values=Entrate_Uscite,
                               textvariable=entrate_uscite_stringvar)
# my_combo.current(0)
my_combo_update.grid(row=5, column=1)

# Bind the ComboBox
my_combo_update.bind("<<ComboboxSelected>>", pick_Categoria_update)

# Categoria ComboBox
categoria_combo_update = ttk.Combobox(Frame_combobox_ok, font=("Helvetica", 10), values=[""],
                                      textvariable=categoria_stringvar)
categoria_combo_update.current(0)
categoria_combo_update.grid(row=6, column=1)

# Bind the ComboBox
categoria_combo_update.bind("<<ComboboxSelected>>", pick_Voce_update)

# Voce Entrata_Spesa ComboBox Combo Box
voce_combo_update = ttk.Combobox(Frame_combobox_ok, font=("Helvetica", 10), values=[""], textvariable=voce_stringvar)
voce_combo_update.current(0)
voce_combo_update.grid(row=7, column=1)

# euro ENTRY
euro_update = Entry(Frame_combobox_ok, font=("Helvetica", 10, 'bold'), bd=5, relief=GROOVE, textvariable=euro_stringvar)
euro_update.grid(row=8, column=1)

####

# class Complex:
#     def __init__(self, realpart, imagpart):
#         self.r = realpart
#         self.i = imagpart
# x = Complex(3.0, -4.5)
# x.r, x.i

anno_report_Intvar = IntVar()

Anno_entry_excell = Entry(Frame_excell, font=('verdana', 10, 'bold'), bg='blue', fg='white',
                          textvariable=anno_report_Intvar)
Anno_entry_excell.grid(row=4, column=1)


# global anno_report

#################################################################
#########################    Class    ###########################
#################################################################

class Report():
    # if __name__ == '__main__':

    def __int__(self, anno):
        self.anno_report = anno


    def anno_report_func(self):

         try:
            self.anno_report = anno_report_Intvar.get()
            return int(self.anno_report)  # return: altrimenti restituisce None - int: altrimenti Type STRING
         except:
             anno_report_Intvar.set(0)
             messagebox.showwarning(title='Dati Mancanti o Errati', message="Scrivere l'anno di interesse")
        # finally:
        #     quit()
    # B_report = Button(Frame_excell_botton, text='report', width=10, command= lambda: print(Report.anno_report_func(anno_report_Stringvar))).grid(row=0, column=2, padx=20, pady=15)

    def report(self):
      try:

        import pandas as pd

        import numpy as np

        import sqlite3

        import openpyxl
        from openpyxl import Workbook
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, Alignment
        from openpyxl.styles import Side, Border
        from openpyxl.styles import PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.worksheet import Worksheet

        import os, sys, subprocess

        # CONNESSIONE A SQLITE3
        conn = sqlite3.connect('database_conti')
        cur = conn.cursor()

        # MI ASSIOCURO DI ESSERE CONNESSO
        # print(conn)
        # print('Sei connesso al database_conti')

        # CREO DATAFRAME df_database_conti
        df_database_conti = pd.read_sql("select * from TABLE_Conti", conn)
        # print(df_database_conti.info(verbose=True))

        # COMMIT e CLOSE
        conn.commit()
        conn.close()

        # LIST TITOLI DATAFRAMES 12 MESI
        list_df_conti_mese_entrate = ['df_database_conti_entrate_gennaio',
                                      'df_database_conti_entrate_febbraio',
                                      'df_database_conti_entrate_marzo',
                                      'df_database_conti_aprile',
                                      'df_database_conti_maggio',
                                      'df_database_conti_giugno',
                                      'df_database_conti_luglio',
                                      'df_database_conti_agosto',
                                      'df_database_conti_settembre',
                                      'df_database_conti_ottobre',
                                      'df_database_conti_novembre',
                                      'df_database_conti_dicembre'
                                      ]

        list_df_conti_mese_uscite = ['df_database_conti_uscite_gennaio',
                                     'df_database_conti_uscite_febbraio',
                                     'df_database_conti_uscite_marzo',
                                     'df_database_conti_uscite_aprile',
                                     'df_database_conti_uscite_maggio',
                                     'df_database_conti_uscite_giugno',
                                     'df_database_conti_uscite_luglio',
                                     'df_database_conti_uscite_agosto',
                                     'df_database_conti_uscite_settembre',
                                     'df_database_conti_uscite_ottobre',
                                     'df_database_conti_uscite_novembre',
                                     'df_database_conti_uscite_dicembre'
                                     ]

        list_mese = ['gennaio',
                     'febbraio',
                     'marzo',
                     'aprile',
                     'maggio',
                     'giugno',
                     'luglio',
                     'agosto',
                     'settembre',
                     'ottobre',
                     'novembre',
                     'dicembre'
                     ]

        ########### imposto anno ##############
        # anno = Report.anno_report_func(anno_report_Stringvar)
        # B_report = Button(Frame_excell_botton, text='report', width=10, command= lambda: print(Report.anno_report_func(anno_report_Stringvar))).grid(row=0, column=2, padx=20, pady=15)

        i = 0
        for x in range(12):
            list_df_conti_mese_entrate[i] = df_database_conti.loc[
                (df_database_conti['Anno'] == Report.anno_report_func(anno_report_Intvar)) &
                (df_database_conti['Mese'] == list_mese[i]) &
                (df_database_conti['Entrate_Uscite'] == 'Entrate')]
            # print(list_df_conti_camerino_pivot_entrate[i].head())

            list_df_conti_mese_uscite[i] = df_database_conti.loc[
                (df_database_conti['Anno'] == Report.anno_report_func(anno_report_Intvar)) &
                (df_database_conti['Mese'] == list_mese[i]) &
                (df_database_conti['Entrate_Uscite'] == 'Uscite')]
            # print(list_df_conti_camerino_mese_uscite[i].head())

            i += 1

        i = 0

        for x in range(12):

            dataframe_empty_list = [(i, Report.anno_report_func(anno_report_Intvar), list_mese[i], 'Entrate', 'vuoto', 'vuoto', 0)]
            if list_df_conti_mese_entrate[i].empty:
                list_df_conti_mese_entrate[i] = pd.DataFrame \
                    (dataframe_empty_list,
                     columns=['index', 'Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro'])

            dataframe_empty_list = [(i, Report.anno_report_func(anno_report_Intvar), list_mese[i], 'Uscite', 'vuoto', 'vuoto', 0)]
            if list_df_conti_mese_uscite[i].empty:
                list_df_conti_mese_uscite[i] = pd.DataFrame \
                    (dataframe_empty_list,
                     columns=['index', 'Anno', 'Mese', 'Entrate_Uscite', 'Categoria', 'Voce', 'Euro'])

            print(list_df_conti_mese_entrate[i].to_markdown())
            print('')
            print(list_df_conti_mese_uscite[i].to_markdown())
            print('')

            # print(list_df_conti_mese_uscite[i].info())
            i += 1

        list_df_conti_mese_entrate = [list_df_conti_mese_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
                                      list_df_conti_mese_entrate[1],
                                      list_df_conti_mese_entrate[2],
                                      list_df_conti_mese_entrate[3],
                                      list_df_conti_mese_entrate[4],
                                      list_df_conti_mese_entrate[5],
                                      list_df_conti_mese_entrate[6],
                                      list_df_conti_mese_entrate[7],
                                      list_df_conti_mese_entrate[8],
                                      list_df_conti_mese_entrate[9],
                                      list_df_conti_mese_entrate[10],
                                      list_df_conti_mese_entrate[11]
                                      ]

        list_df_conti_mese_uscite = [list_df_conti_mese_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
                                     list_df_conti_mese_uscite[1],
                                     list_df_conti_mese_uscite[2],
                                     list_df_conti_mese_uscite[3],
                                     list_df_conti_mese_uscite[4],
                                     list_df_conti_mese_uscite[5],
                                     list_df_conti_mese_uscite[6],
                                     list_df_conti_mese_uscite[7],
                                     list_df_conti_mese_uscite[8],
                                     list_df_conti_mese_uscite[9],
                                     list_df_conti_mese_uscite[10],
                                     list_df_conti_mese_uscite[11]
                                     ]

        list_df_conti_mese_entrate = [list_df_conti_mese_entrate[0],  # df_conti_camerino_pivot_entrate_gennaio
                                      list_df_conti_mese_entrate[1],
                                      list_df_conti_mese_entrate[2],
                                      list_df_conti_mese_entrate[3],
                                      list_df_conti_mese_entrate[4],
                                      list_df_conti_mese_entrate[5],
                                      list_df_conti_mese_entrate[6],
                                      list_df_conti_mese_entrate[7],
                                      list_df_conti_mese_entrate[8],
                                      list_df_conti_mese_entrate[9],
                                      list_df_conti_mese_entrate[10],
                                      list_df_conti_mese_entrate[11]
                                      ]

        list_df_conti_mese_uscite = [list_df_conti_mese_uscite[0],  # df_conti_camerino_pivot_uscite_gennaio
                                     list_df_conti_mese_uscite[1],
                                     list_df_conti_mese_uscite[2],
                                     list_df_conti_mese_uscite[3],
                                     list_df_conti_mese_uscite[4],
                                     list_df_conti_mese_uscite[5],
                                     list_df_conti_mese_uscite[6],
                                     list_df_conti_mese_uscite[7],
                                     list_df_conti_mese_uscite[8],
                                     list_df_conti_mese_uscite[9],
                                     list_df_conti_mese_uscite[10],
                                     list_df_conti_mese_uscite[11]
                                     ]

        # print(list_df_conti_mese_entrate[0].columns)

        def pivot_table_w_subtotals(df, values, indices, columns, aggfunc, fill_value):
            listOfTable = []
            for indexNumber in range(len(indices)):
                n = indexNumber + 1
                if n == 1:
                    table = pd.pivot_table(df, values=values, index=indices[:n], columns=columns, aggfunc=aggfunc,
                                           fill_value=fill_value)

                else:
                    table = pd.pivot_table(df, values=values, index=indices[:n], columns=columns, aggfunc=aggfunc,
                                           fill_value=fill_value)
                table = table.reset_index()

                for column in indices[n:]:
                    table[column] = ''

                listOfTable.append(table)

            concatTable = pd.concat(listOfTable).sort_index()
            concatTable = concatTable.set_index(keys=indices)
            return concatTable.sort_index(axis=0, ascending=True)

        list_pivot_mese_entrate = ['pivot_gennaio_entrate',
                                   'pivot_febbraio_entrate',
                                   'pivot_marzo_entrate',
                                   'pivot_aprile_entrate',
                                   'pivot_maggio_entrate',
                                   'pivot_giugno_entrate',
                                   'pivot_luglio_entrate',
                                   'pivot_agosto_entrate',
                                   'pivot_settembre_entrate',
                                   'pivot_ottobre_entrate',
                                   'pivot_novembre_entrate',
                                   'pivot_dicembre_entrate'
                                   ]

        list_pivot_mese_uscite = ['pivot_gennaio_uscite',
                                  'pivot_febbraio_uscite',
                                  'pivot_marzo_uscite',
                                  'pivot_aprile_uscite',
                                  'pivot_maggio_uscite',
                                  'pivot_giugno_uscite',
                                  'pivot_luglio_uscite',
                                  'pivot_agosto_uscite',
                                  'pivot_settembre_uscite',
                                  'pivot_ottobre_uscite',
                                  'pivot_novembre_uscite',
                                  'pivot_dicembre_uscite'
                                  ]

        # pivot_table_w_subtotals
        #       (df=list_df_conti_camerino_mese_entrate[0],values='Euro',indices=['Entrate_Uscite', 'Categoria', 'Voce'],columns=[],aggfunc='sum',fill_value='')

        i = 0
        try:
            for x in range(12):
                list_pivot_mese_entrate[i] = np.round(pivot_table_w_subtotals
                                                  (df=list_df_conti_mese_entrate[i],
                                                   values='Euro',
                                                   indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                   columns=[],
                                                   aggfunc='sum',
                                                   fill_value=''), 2)

                list_pivot_mese_uscite[i] = np.round(pivot_table_w_subtotals
                                                 (list_df_conti_mese_uscite[i],
                                                  values='Euro',
                                                  indices=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                  columns=[],
                                                  aggfunc='sum',
                                                  fill_value=''), 2)
                i += 1
        except:
            pass



        list_pivot_mese_entrate = [list_pivot_mese_entrate[0],  # pivot gennaio entrate
                                   list_pivot_mese_entrate[1],
                                   list_pivot_mese_entrate[2],
                                   list_pivot_mese_entrate[3],
                                   list_pivot_mese_entrate[4],
                                   list_pivot_mese_entrate[5],
                                   list_pivot_mese_entrate[6],
                                   list_pivot_mese_entrate[7],
                                   list_pivot_mese_entrate[8],
                                   list_pivot_mese_entrate[9],
                                   list_pivot_mese_entrate[10],
                                   list_pivot_mese_entrate[11]
                                   ]

        list_pivot_mese_uscite = [list_pivot_mese_uscite[0],
                                  list_pivot_mese_uscite[1],
                                  list_pivot_mese_uscite[2],
                                  list_pivot_mese_uscite[3],
                                  list_pivot_mese_uscite[4],
                                  list_pivot_mese_uscite[5],
                                  list_pivot_mese_uscite[6],
                                  list_pivot_mese_uscite[7],
                                  list_pivot_mese_uscite[8],
                                  list_pivot_mese_uscite[9],
                                  list_pivot_mese_uscite[10],
                                  list_pivot_mese_uscite[11]
                                  ]

        # print(list_pivot_mese_entrate[0])

        ##############PRIMA PAGINA#############################

        # Creo il file 'conti_styled.xlsx'
        wb = Workbook()
        # La prima pagina 'Sheet' la chiamo 'Copertina_fronte'
        wb['Sheet'].title = ('Copertina_fronte')

        wb['Copertina_fronte'].merge_cells('A4:I4')
        wb['Copertina_fronte']['A4'] = 'Resoconto Amministrativo'
        wb['Copertina_fronte']['A4'].font = Font(name='Calibri',
                                                 size=35,
                                                 bold=True,
                                                 italic=True,
                                                 vertAlign='none',
                                                 underline='single',
                                                 strike=False,
                                                 color='204ac8')  # blu royal
        wb['Copertina_fronte']['A4'].alignment = Alignment(horizontal='center')

        wb['Copertina_fronte']['A7'] = 'Fraternità di .....'
        wb['Copertina_fronte']['A7'].font = Font(name='Calibri',
                                                 size=25,
                                                 bold=True,
                                                 italic=True,
                                                 vertAlign='none',
                                                 underline='single',
                                                 strike=False,
                                                 color='204ac8')

        wb['Copertina_fronte']['A10'] = 'Anno.....'
        wb['Copertina_fronte']['A10'].font = Font(name='Calibri',
                                                  size=25,
                                                  bold=True,
                                                  italic=True,
                                                  vertAlign='none',
                                                  underline='single',
                                                  strike=False,
                                                  color='204ac8')

        # Inserisco immagine bilancia
        img = openpyxl.drawing.image.Image('bilancia.png')
        img.anchor = 'B13'
        wb['Copertina_fronte'].add_image(img)

        wb.save('conti_styled.xlsx')
        ##################################

        # Con ExcelWriter di pandas METTO INSIEME il pivot delle entrate e il pivot delle uscite

        # sheets dei 12 mesi
        list_ws_mese = ['ws_gennaio',
                        'ws_febbraio',
                        'ws_marzo',
                        'ws_aprile',
                        'ws_maggio',
                        'ws_giugno',
                        'ws_luglio',
                        'ws_agosto',
                        'ws_settembre',
                        'ws_ottobre',
                        'ws_novembre',
                        'ws_dicembre'
                        ]

        c = 0  # contatore
        try:
            for x in range(12):
                with pd.ExcelWriter('conti_styled.xlsx',
                                    mode="a",
                                    engine="openpyxl",
                                    if_sheet_exists="overlay",
                                    ) as writer:
                    list_pivot_mese_entrate[c].to_excel(writer, sheet_name=list_mese[c], startrow=5)
                    list_pivot_mese_uscite[c].to_excel(writer, sheet_name=list_mese[c],
                                                    startrow=(len(list_pivot_mese_entrate[c]) + 10))

                # leggo il file "conti_camerino_styled.xlsx"
                wb = load_workbook(filename="conti_styled.xlsx")
                # creo i 12 sheet
                list_ws_mese[c] = wb[list_mese[c]]
                c += 1
        except:
            pass


        # sheets dei 12 mesi
        list_ws_mese = [wb[list_mese[0]],  # ws_gennaio,
                        wb[list_mese[1]],  # ws_febbraio,
                        wb[list_mese[2]],  # ws_marzo,
                        wb[list_mese[3]],  # ws_aprile,
                        wb[list_mese[4]],  # ws_maggio,
                        wb[list_mese[5]],  # ws_giugno,
                        wb[list_mese[6]],  # ws_luglio,
                        wb[list_mese[7]],  # ws_agosto,
                        wb[list_mese[8]],  # ws_settembre,
                        wb[list_mese[9]],  # ws_ottobre,
                        wb[list_mese[10]],  # ws_novembre,
                        wb[list_mese[11]],  # ws_dicembre
                        ]

        ################# APPLICO STILE ########################

        # Colonna D :Formattazione degli euro in valuta euro
        for sheet in list_ws_mese:
            for row in sheet[7:sheet.max_row]:  # skip the header
                # print(row) #(<Cell 'gennaio'.A7>, <Cell 'gennaio'.B7>, <Cell 'gennaio'.C7>, <Cell 'gennaio'.D7>)
                cell = row[3]  # il quarto valore della tuple
                print(cell)  # <Cell 'multiple'.D7>
                cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(bold=True)

        # Aggiungo la scritta 'Totale =' alla tabella pivot davanti ai subtotali
        for sheet in list_ws_mese:
            for row in sheet[7:sheet.max_row]:
                for cell in sheet['B']:  # per ogni casella della colonna B
                    if cell.value is not None:
                        # ossia se la casella nella colonna B non è vuota
                        # Assegnale uno stile
                        cell.font = Font(name='Calibri',
                                         size=15,
                                         bold=True,
                                         italic=True,
                                         vertAlign='none',
                                         underline='single',
                                         strike=False,
                                         color='a81a1a')
                        # Assegna uno stile anche alla cell accanto corrispondente
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=3,
                                   value=f"Totale {cell.value} =")
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=0, column=0).row, column=4).alignment \
                            = Alignment(horizontal="left")

        # Larghezza fissa colonne
        i = 0  # contatore
        # set the height of the first row in each sheet
        for sheet in list_ws_mese:
            sheet.row_dimensions[1].height = 70

            # set the width of the column
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 30
            sheet.column_dimensions['D'].width = 15

            # merge cells
            sheet.merge_cells('A1:D1')

            # scrivo nella cella 'A1'
            sheet['A1'].value = list_mese[i]
            i += 1

            # Formattazione cella
            sheet['A1'].font = Font(name='Calibri',
                                    size=25,
                                    bold=True,
                                    italic=True,
                                    vertAlign='none',
                                    underline='single',
                                    strike=False,
                                    color='a81a1a')

            sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

            # Colonna C: Allineamento
            for row in sheet[7:sheet.max_row]:  # skip the header
                cell = row[2]  # il terzo valore della tuple
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Colonna D: Allineamento
            for row in sheet[7:sheet.max_row]:  # skip the header
                cell = row[1]  # il secondo valore della tuple
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formattazione headers
            list = []

            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ("Categoria") or
                            cell.value == ("Entrate") or
                            cell.value == ("Euro") or
                            cell.value == ("Uscite") or
                            cell.value == ("Voce")):
                        list.append(cell)
            for cell in list:
                cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ("Totale Categoria =")):
                        cell.font = Font(name='Calibri', size=15, color='a81a1a', bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")

            # Formattazione ' Euro accanto a 'Totale categoria='

            double = Side(border_style="double", color="4617F1")
            cont = 0
            for row in sheet.rows:
                for cell in row:
                    if (cell.value == ('Totale Categoria =')):
                        #
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).alignment \
                            = Alignment(horizontal="left")
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).fill \
                            = PatternFill('solid', fgColor='d1d22e')
                        sheet.cell(row=cell.offset(row=0, column=1).row, column=4).border \
                            = Border(bottom=double, top=double, left=double, right=double)

                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).font \
                            = Font(size=15, color='a81a1a', bold=True)
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).number_format \
                            = '#,##0.00 €'
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).alignment \
                            = Alignment(horizontal="left")
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).fill \
                            = PatternFill('solid', fgColor='d1d22e')
                        sheet.cell(row=cell.offset(row=1, column=1).row, column=4).border \
                            = Border(bottom=double, top=double, left=double, right=double)

            # Rendi 'invisibile il testo"Entrate_Uscite"
            list = []

            # Formattazione "Entrate_Uscite", piccolo per non essere visto

            for row in sheet.rows:
                for cell in row:
                    if cell.value == ("Entrate_Uscite"):
                        list.append(cell)
            for cell in list:
                cell.font = Font(size=1)

        # imposto saldo iniziale
        saldo = 200_000
        # imposto un contatore
        e = 0
        i = 0
        # memorizzo in una lista tutti i saldi dall'inizio alla fine dell'anno
        # le entrete
        # le uscite
        list_saldo_iniale_finale_anno = [saldo]
        list_entrate_mesi = []
        list_uscite_mesi = []
        # saldo iniziale
        for sheet in list_ws_mese:
            # coordinate dell'ultima cella della colonna A di ogni foglio

            last_cell_coordiate = 'C' + str(sheet.max_row)
            # print(last_cell_coordiate)
            # attraverso le coordinate risalgo alla cella di excel
            cell = sheet[last_cell_coordiate]
            # print(cell.value)

            # cell = sheet.cell(row=1, column=1)
            # last_cell = sheet[last_cell]
            #
            # print(cell.coordinate, cell.row, cell.column) # A184 184 1 per tutti e 12 i fogli
            #
            # #print(cell) # stampa ultima cella colonna A
            sheet.cell(row=cell.offset(row=5, column=0).row, column=2, value='SALDO del mese precedente').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=7, column=0).row, column=2, value='ENTRATE del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=7, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=9, column=0).row, column=2, value='USCITE del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=9, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=11, column=0).row, column=2, value='DIS/AVANZO del mese').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=11, column=0).row, column=2).alignment = Alignment(horizontal="left")
            sheet.cell(row=cell.offset(row=13, column=0).row, column=2, value='SALDO del mese corrente').font \
                = Font(size=15, color='000000', bold=True)
            sheet.cell(row=cell.offset(row=13, column=0).row, column=2).alignment = Alignment(horizontal="left")

            sheet.cell(row=cell.offset(row=5, column=0).row, column=4,
                       value=(
                           saldo
                       )
                       )

            # mi calcolo il saldo finale e la assegno alla variabile saldo
            saldo = (saldo +
                     list_df_conti_mese_entrate[e]['Euro'].sum(numeric_only=True) -
                     list_df_conti_mese_uscite[e]['Euro'].sum(numeric_only=True)
                     )

            # print(list_saldo_iniale_finale_anno)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(size=15, bold=True)
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=5, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if sheet.cell(row=cell.offset(row=5, column=0).row, column=4).value > 0:
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                       bold=True)
            else:
                sheet.cell(row=cell.offset(row=5, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                       bold=True)

            sheet.cell(row=cell.offset(row=7, column=0).row, column=4,
                       value=list_df_conti_mese_entrate[i]['Euro'].sum(numeric_only=True))

            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).font = Font(size=15, color='000000',
                                                                                   bold=True)
            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=7, column=0).row, column=4).alignment = Alignment(horizontal="right")
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4,
                       value=list_df_conti_mese_uscite[i]['Euro'].sum(numeric_only=True))
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).font = Font(size=15, color='a81a1a',
                                                                                   bold=True)
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).number_format = '-#,##0.00€'
            sheet.cell(row=cell.offset(row=9, column=0).row, column=4).alignment = Alignment(horizontal="right")
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4,
                       value=((list_df_conti_mese_entrate[i]['Euro']).sum(numeric_only=True) -
                              (list_df_conti_mese_uscite[i]['Euro']).sum(numeric_only=True)
                              )
                       )
            # sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(size=15,
            #                                                                       bold=True)
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=11, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if (sheet.cell(row=cell.offset(row=11, column=0).row, column=4).value) > 0:
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                        bold=True)
            else:
                sheet.cell(row=cell.offset(row=11, column=0).row, column=4).font = Font(color='a81a1a', size=15,

                                                                                        bold=True)
            # queste liste mi servono per il grafico
            list_saldo_iniale_finale_anno.append(saldo)
            list_entrate_mesi.append(list_df_conti_mese_entrate[i]['Euro'].sum(numeric_only=True))
            list_uscite_mesi.append(list_df_conti_mese_uscite[i]['Euro'].sum(numeric_only=True))

            # print(list_saldo_iniale_finale_anno)
            # print(list_entrate_mesi)
            # print(list_uscite_mesi)

            i += 1
            # Saldo finale
            # for row in sheet:
            #         for cell in row:
            #             if (cell.value == ("TOTALE_Uscite")):

            sheet.cell(row=cell.offset(row=13, column=0).row, column=4,
                       value=(
                           saldo
                       )
                       )

            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(size=15, bold=True)
            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).number_format = '#,##0.00€'
            sheet.cell(row=cell.offset(row=13, column=0).row, column=4).alignment = Alignment(horizontal="right")

            if sheet.cell(row=cell.offset(row=13, column=0).row, column=4).value > 0:
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='000000', size=15,
                                                                                        bold=True)
            else:
                sheet.cell(row=cell.offset(row=13, column=0).row, column=4).font = Font(color='a81a1a', size=15,
                                                                                        bold=True)

        ###################### tabellone entrate
        import openpyxl
        from openpyxl.worksheet import page

        # Creo un nuovo foglio
        ws_tab_entrate = wb.create_sheet('Tab_Entrate')
        ws_tab_entrate.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)
        # ws_tab_entrate.print_area = 'A1:Z1'

        ws_tab_entrate['A1'] = 'Tabellone Entrate'
        ws_tab_entrate.row_dimensions[1].height = 70
        ws_tab_entrate['A1'].font = Font(name='Calibri', size=80, color='a81a1a', bold=True)
        ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_tab_entrate.merge_cells('A1:N1')

        # set the width of the column
        ws_tab_entrate.column_dimensions['A'].width = 14
        ws_tab_entrate.column_dimensions['B'].width = 8
        ws_tab_entrate.column_dimensions['C'].width = 8
        ws_tab_entrate.column_dimensions['D'].width = 8
        ws_tab_entrate.column_dimensions['E'].width = 8
        ws_tab_entrate.column_dimensions['F'].width = 8
        ws_tab_entrate.column_dimensions['G'].width = 8
        ws_tab_entrate.column_dimensions['H'].width = 8
        ws_tab_entrate.column_dimensions['I'].width = 8
        ws_tab_entrate.column_dimensions['L'].width = 9
        ws_tab_entrate.column_dimensions['M'].width = 8
        ws_tab_entrate.column_dimensions['N'].width = 14

        i = 2
        for x in range(2, 20):
            ws_tab_entrate.row_dimensions[i].height = 20
            i += 1

        # anno =
        df_conti_camerino_TOT_entrate = df_database_conti.loc[
            (df_database_conti['Anno'] == Report.anno_report_func(anno_report_Intvar)) &
            (df_database_conti['Entrate_Uscite'] == 'Entrate')]
        # print(df_conti_camerino_TOT_entrate.head(40))

        pivot_conti_camerino_TOT_entrate = np.round(pd.pivot_table
                                                    (df_conti_camerino_TOT_entrate,
                                                     values='Euro',
                                                     index=['Categoria'],
                                                     columns='Mese',
                                                     aggfunc='sum',
                                                     margins=True,
                                                     margins_name='TOTALE_Entrate',
                                                     fill_value=0), 2)

        # print(pivot_conti_camerino_TOT_entrate.head())

        for r in dataframe_to_rows(pivot_conti_camerino_TOT_entrate, index=True, header=True):
            ws_tab_entrate.append(r)

        for cell in ws_tab_entrate['A'] + ws_tab_entrate[2]:
            cell.style = 'Pandas'

        ws_tab_entrate['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
        ws_tab_entrate['A1'].alignment = Alignment(horizontal="center", vertical="center")

        TOTALE_ENTRATE = round(df_conti_camerino_TOT_entrate['Euro'].sum(), 2)

        for row in ws_tab_entrate.rows:
            for cell in row:
                if cell.value == TOTALE_ENTRATE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,##0.00€'

            # list_df_conti_camerino_mese_uscite[i] = df_conti_camerino_modified.loc[
            #     (df_conti_camerino_modified['Anno'] == anno) &
            #     (df_conti_camerino_modified['Mese'] == list_mese[i]) &
            #     (df_conti_camerino_modified['Entrate_Uscite'] == 'Uscite')]
            # #print(list_df_conti_camerino_mese_uscite[i].head())
            #
            # i += 1

        ###################### tabellone uscite
        # Creo un nuovo foglio
        ws_tab_uscite = wb.create_sheet('Tab_Uscite')

        ws_tab_uscite.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)

        ws_tab_uscite['A1'] = 'Tabellone Uscite'
        ws_tab_uscite.merge_cells('A1:N1')
        ws_tab_uscite.row_dimensions[1].height = 45

        df_conti_camerino_TOT_uscite = df_database_conti.loc[
            (df_database_conti['Anno'] == Report.anno_report_func(anno_report_Intvar)) &
            (df_database_conti['Entrate_Uscite'] == 'Uscite')]

        # print(df_conti_camerino_TOT_uscite.head(40))
        pivot_conti_camerino_TOT_uscite = np.round(pd.pivot_table
                                                   (df_conti_camerino_TOT_uscite,
                                                    values='Euro',
                                                    # index=['Entrate_Uscite', 'Categoria', 'Voce'],
                                                    index=['Categoria'],
                                                    columns='Mese',
                                                    aggfunc='sum',
                                                    margins=True,
                                                    margins_name='TOTALE_Uscite',
                                                    fill_value=0), 2)

        print(pivot_conti_camerino_TOT_uscite.head())
        for r in dataframe_to_rows(pivot_conti_camerino_TOT_uscite, index=True, header=True):
            ws_tab_uscite.append(r)

        for cell in ws_tab_uscite['A'] + ws_tab_uscite[2]:
            cell.style = 'Pandas'

        # set the width of the column
        ws_tab_uscite.column_dimensions['A'].width = 16
        ws_tab_uscite.column_dimensions['B'].width = 7
        ws_tab_uscite.column_dimensions['C'].width = 7
        ws_tab_uscite.column_dimensions['D'].width = 7
        ws_tab_uscite.column_dimensions['E'].width = 7
        ws_tab_uscite.column_dimensions['F'].width = 7
        ws_tab_uscite.column_dimensions['G'].width = 7
        ws_tab_uscite.column_dimensions['H'].width = 7
        ws_tab_uscite.column_dimensions['I'].width = 7
        ws_tab_uscite.column_dimensions['L'].width = 9
        ws_tab_uscite.column_dimensions['M'].width = 7
        ws_tab_uscite.column_dimensions['N'].width = 16

        i = 2
        for x in range(2, 20):
            ws_tab_uscite.row_dimensions[i].height = 17
            i += 1

        ws_tab_uscite['A1'].font = Font(name='Calibri', size=40, color='a81a1a', bold=True)
        ws_tab_uscite['A1'].alignment = Alignment(horizontal="center", vertical="center")

        TOTALE_USCITE = round(df_conti_camerino_TOT_uscite['Euro'].sum(), 2)

        for row in ws_tab_uscite.rows:
            for cell in row:
                if cell.value == TOTALE_USCITE:
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = Border(bottom=double, top=double, left=double, right=double)
                    cell.fill = PatternFill('solid', fgColor='d1d22e')
                    cell.number_format = '#,##0.00€'

        ######################  grafico

        from openpyxl.chart import Reference, LineChart

        # Creo un nuovo foglio
        ws_saldo_riepilogo = wb.create_sheet('Saldo_riepilogo')
        ws_saldo_riepilogo['A1'] = 'Tabella Entrate - Uscite - Saldo di ogni mese '

        # Formattazione
        ws_saldo_riepilogo.row_dimensions[1].height = 100
        ws_saldo_riepilogo.merge_cells('A1:E1')

        # set the width of the column
        ws_saldo_riepilogo.column_dimensions['A'].width = 15
        ws_saldo_riepilogo.column_dimensions['B'].width = 17
        ws_saldo_riepilogo.column_dimensions['C'].width = 17
        ws_saldo_riepilogo.column_dimensions['D'].width = 17
        ws_saldo_riepilogo.column_dimensions['E'].width = 17

        # queste liste mi servono per il grafico

        list_headers = ['mese', 'saldo_iniziale', 'entrate_mese', 'uscite_mese', 'saldo_finale']
        list_saldo_iniziale_anno = list_saldo_iniale_finale_anno[:-1]
        # list_entrate_mesi
        # list_uscite_mesi
        list_saldo_finale_anno = list_saldo_iniale_finale_anno[1:]

        i = 0
        ws_saldo_riepilogo.append(list_headers)
        for mese in range(1, 13):
            mese_saldo_grafico = [list_mese[i], list_saldo_iniziale_anno[i], list_entrate_mesi[i], list_uscite_mesi[i],
                                  list_saldo_finale_anno[i]]

            ws_saldo_riepilogo.append(mese_saldo_grafico)
            i += 1

        list = []
        for row in ws_saldo_riepilogo.rows:
            for cell in row:

                if (cell.value == ("saldo_iniziale") or
                        cell.value == ("entrate_mese") or
                        cell.value == ("uscite_mese") or
                        cell.value == ("saldo_finale") or
                        cell.value == ("gennaio") or
                        cell.value == ("febbraio") or
                        cell.value == ("marzo") or
                        cell.value == ("aprile") or
                        cell.value == ("maggio") or
                        cell.value == ("giugno") or
                        cell.value == ("luglio") or
                        cell.value == ("agosto") or
                        cell.value == ("settembre") or
                        cell.value == ("ottobre") or
                        cell.value == ("novembre") or
                        cell.value == ("dicembre")
                ):
                    cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for cell in row:
                if cell.value == ("mese"):
                    cell.font = Font(size=1)

        for row in ws_saldo_riepilogo.iter_rows(min_row=3, min_col=2, max_row=14, max_col=5):
            for cell in row:
                cell.font = Font(name='Calibri', size=13, color='000000', bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 €'
                if (int(cell.value) > 0):
                    cell.font = Font(color='000000')
                else:
                    cell.font = Font(color='a81a1a')

        ws_saldo_riepilogo['A1'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
        ws_saldo_riepilogo['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_saldo_riepilogo['B3'].border \
            = Border(bottom=double, top=double, left=double, right=double)
        ws_saldo_riepilogo['B3'].fill \
            = PatternFill('solid', fgColor='d1d22e')
        ws_saldo_riepilogo['E14'].border \
            = Border(bottom=double, top=double, left=double, right=double)
        ws_saldo_riepilogo['E14'].fill \
            = PatternFill('solid', fgColor='d1d22e')

        data = Reference(ws_saldo_riepilogo, min_col=3, min_row=2, max_col=5, max_row=14)
        titles = Reference(ws_saldo_riepilogo, min_row=3, max_row=14, min_col=1)

        chart = LineChart()
        chart.title = "Bilancio"
        chart.style = 12

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = 'Mesi'
        chart.y_axis.title = 'Euro'

        ws_saldo_riepilogo.add_chart(chart, "A21")
        #######################PAGINA CONCLUSIVA
        # format
        from openpyxl.styles import NamedStyle, Font, Border, Side
        # HIGHLIGHT
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(name='Calibri', size=15, color='000000', bold=True)
        double = Side(border_style="double", color="4617F1")
        # highlight.border = Border(bottom=double, top=double, left=double, right=double)
        highlight.fill = PatternFill('solid', fgColor='d1d22e')
        highlight.alignment = Alignment(horizontal="right", vertical="center")
        highlight.number_format = '#,##0.00 €'
        wb.add_named_style(highlight)

        ############BLACK
        black = NamedStyle(name="black")
        black.font = Font(name='Calibri', size=15, color='000000', bold=True)
        black.alignment = Alignment(horizontal="right", vertical="center")
        wb.add_named_style(black)

        import openpyxl
        from openpyxl import load_workbook
        ws_fine = wb.create_sheet('Fine')
        ws_fine['A1'] = 'Bilancio Anno ...'
        ws_fine.merge_cells('A1:I1')
        ws_fine.row_dimensions[1].height = 45
        ws_fine['A1'].font = Font(name='Calibri', size=35, color='a81a1a', bold=True)
        ws_fine['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws_fine['A7'] = 'RIEPILOGO'
        ws_fine.merge_cells('A7:C7')

        ws_fine['A10'] = 'SALDO iniziale'
        ws_fine.merge_cells('A10:C10')
        ws_fine['E10'] = int(200_000)
        ws_fine.merge_cells('E10:G10')
        ws_fine['E10'].style = 'highlight'

        ws_fine['A13'] = 'TOTALE Entrate'
        ws_fine.merge_cells('A13:C13')
        ws_fine['E13'] = TOTALE_ENTRATE
        ws_fine.merge_cells('E13:G13')
        ws_fine['E13'].style = 'highlight'

        ws_fine['A16'] = 'TOTALE Uscite'
        ws_fine.merge_cells('A16:C16')
        ws_fine['E16'] = TOTALE_USCITE
        ws_fine.merge_cells('E16:G16')
        ws_fine['E16'].style = 'highlight'

        ws_fine['A19'] = 'DIS/AVANZO'
        ws_fine.merge_cells('A19:C19')
        ws_fine['E19'] = (TOTALE_ENTRATE - TOTALE_USCITE)
        ws_fine.merge_cells('E19:G19')
        ws_fine['E19'].style = 'highlight'

        ws_fine['A22'] = 'SALDO Finale'
        ws_fine.merge_cells('A22:C22')
        ws_fine['E22'] = saldo
        ws_fine.merge_cells('E22:G22')
        ws_fine['E22'].style = 'highlight'

        ws_fine['A7'].font = Font(name='Calibri', size=20, color='a81a1a', bold=True)
        ws_fine['A7'].alignment = Alignment(horizontal="center", vertical="center")
        ws_fine['A10'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A10'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A13'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A13'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A16'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A16'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A19'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A19'].alignment = Alignment(horizontal="right", vertical="center")
        ws_fine['A22'].font = Font(name='Calibri', size=15, bold=True)
        ws_fine['A22'].alignment = Alignment(horizontal="right", vertical="center")

        ws_fine['A32'] = 'Data'
        ws_fine['A32'].style = 'black'
        ws_fine.merge_cells('A32:B32')
        ws_fine['D29'] = 'Guardiano'
        ws_fine['D29'].style = 'black'
        ws_fine.merge_cells('D29:I29')
        ws_fine['d35'] = 'Vicario'
        ws_fine['D35'].style = 'black'
        ws_fine.merge_cells('D35:I35')
        ws_fine['D41'] = 'Economo'
        ws_fine['D41'].style = 'black'
        ws_fine.merge_cells('D41:I41')
        ws_fine['A38'] = 'Timbro'
        ws_fine['A38'].style = 'black'
        ws_fine.merge_cells('A38:B38')



        wb.save('conti_styled.xlsx')

        if sys.platform == "win32":
            os.startfile('conti_styled.xlsx')
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, 'conti_styled.xlsx'])
      except:
            messagebox.showwarning(title='Dati Mancanti', message="L'anno indicato deve avere almeno una voce in entrata e una voce in uscita")







# B_add = Button(Frame1in, text='add', width=10, command=lambda:[submit()]).grid(row=0, column=0, padx=20, pady=15)

B_add = Button(Frame1in, text='aggiungi', width=10, command=lambda: [submit(), query_database()]).grid(row=0, column=0,
                                                                                                       padx=390,
                                                                                                       pady=15)
B_update = Button(Frame_update_botton, text='aggiorna', width=10, command=update_record).grid(row=0, column=1, padx=20,
                                                                                              pady=15)
B_delete = Button(Frame_update_botton, text='cancella', width=10, command=remove_one).grid(row=0, column=2, padx=20,
                                                                                           pady=15)
B_excel = Button(Frame_excell_botton, text='excel', width=10, command=sqlite3_to_excel).grid(row=0, column=1, padx=20,
                                                                                             pady=15)
# B_report = Button(Frame_excell_botton, text='report', width=10, command= lambda: print(Report.anno_report_func(anno_report_Stringvar))).grid(row=0, column=2, padx=20, pady=15)
B_report = Button(Frame_excell_botton, text='report', width=10,
                  command=lambda: Report.report(Report.anno_report_func(anno_report_Intvar))).grid(row=0, column=2,
                                                                                                      padx=20, pady=15)

# B_report = Button(Frame_excell_botton, text='report', width=10, command=lambda: Report.report(2012)).grid(row=0, column=2, padx=20, pady=15)
B_esporta = Button(Frame_excell_botton, text='esporta', width=10, command='').grid(row=0, column=3, padx=20, pady=15)
B_importa = Button(Frame_excell_botton, text='importa', width=10, command='').grid(row=0, column=4, padx=20, pady=15)
#####


query_database()

# conn.close()

root.mainloop()
